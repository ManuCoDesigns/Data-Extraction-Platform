#!/usr/bin/env python3
"""
Critical Materials Supplier Extractor  v2.1
============================================
Reads Intern-_Critical_Materials_Suppliers.xlsx, fetches each company website,
and emits one BGS Supplier Graph Schema JSON record per supplier.

Design principles
-----------------
• BGS schema fields are mapped/inferred; all non-schema data goes into `extras`.
• Multi-type operators (Mine/Refiner/Producer) get the *lowest* (most upstream)
  tier so the record is as useful as possible for sourcing.
• Products are parsed to one entry per distinct product name + grade combination.
• Manufacturing sites are extracted from capacity data using heuristic patterns.
• The website fetch is optional (--skip-fetch) and fails gracefully.
• Duplicate companies across sheets are deduplicated and their data merged.

Usage
-----
    python extractor.py
    python extractor.py --skip-fetch
    python extractor.py --limit 10 --workers 2 --delay 1.5
    python extractor.py --xlsx path/to/file.xlsx --out output.json
"""

import argparse
import json
import logging
import re
import time
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Optional
from urllib.parse import urljoin, urlparse

import openpyxl
import requests
from bs4 import BeautifulSoup

# Claude LLM (optional — only needed when --use-llm is set)
try:
    import anthropic
    _HAS_ANTHROPIC = True
except ImportError:
    _HAS_ANTHROPIC = False

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# BGS Schema Fixed Values
# ---------------------------------------------------------------------------
# Sources are built dynamically per company so source_url matches the actual
# page scraped (company website), not a hardcoded publication URL.

def make_source(source_name: str, source_url, tier: str = "tier1") -> dict:
    return {"source_name": source_name, "source_url": source_url,
            "doi": None, "tier": tier}

DATA_COMPLETENESS_FLAGS = {
    "review_score": "manual_only",
    "defect_rate_ppm": "manual_only",
    "on_time_delivery_rate": "manual_only",
    "pricing": "api_only",
    "inventory_levels": "api_only",
}

# ---------------------------------------------------------------------------
# Sheet configuration
# ---------------------------------------------------------------------------

# Sheets containing actual supplier rows (skip reference/index sheets)
SUPPLIER_SHEETS = {
    "Other mines- recyclers",
    "Lithium Producers",
    "Rare Earth Elements",
    "Cobalt Producers",
    "Copper Producers",
    "Nickel Producers",
    "Graphite Producers",
    "Aluminum Producers",
    "Zinc Lead Tin Producers",
    "Multi-Metal Refiners",
    "Specialty Critical Materials",
}

# Map sheet name → primary commodity context (used as fallback for sector inference)
SHEET_COMMODITY_CONTEXT = {
    "Lithium Producers": "lithium",
    "Rare Earth Elements": "rare earth",
    "Cobalt Producers": "cobalt",
    "Copper Producers": "copper",
    "Nickel Producers": "nickel",
    "Graphite Producers": "graphite",
    "Aluminum Producers": "aluminum",
    "Zinc Lead Tin Producers": "zinc",
    "Multi-Metal Refiners": "copper",
    "Specialty Critical Materials": "critical minerals",
    "Other mines- recyclers": "recycled materials",
}

# ---------------------------------------------------------------------------
# Tier & Sector mapping tables
# ---------------------------------------------------------------------------

# Lower tier number = more upstream.  When a company has multiple type tokens
# (e.g. Mine/Refiner/Producer), we pick the *minimum* tier (most upstream).
TIER_MAP = {
    "explorer":         1,
    "developer":        1,
    "mine":             1,
    "miner":            1,
    "quarry":           1,
    "pit":              1,
    "producer":         1,
    "mine/producer":    1,
    "mine/processor":   1,   # mining is primary
    "mine/refiner":     1,
    "processor":        2,
    "refiner":          2,
    "smelter":          2,
    "recycler":         2,
    "manufacturer":     2,
    "supplier":         3,
    "trader":           3,
    "distributor":      3,
}

# Keyword → industry_sector (checked in order; first match wins)
SECTOR_MAP = [
    ("recycl",          "recycled aggregates"),
    ("battery recycl",  "recycled aggregates"),
    ("rare earth",      "metals mining"),
    ("ree",             "metals mining"),
    ("ndpr",            "metals mining"),
    ("dysprosium",      "metals mining"),
    ("terbium",         "metals mining"),
    ("lithium",         "metals mining"),
    ("cobalt",          "metals mining"),
    ("nickel",          "metals mining"),
    ("copper",          "metals mining"),
    ("zinc",            "metals mining"),
    ("lead",            "metals mining"),
    ("tin",             "metals mining"),
    ("aluminum",        "metals mining"),
    ("aluminium",       "metals mining"),
    ("bauxite",         "metals mining"),
    ("alumina",         "metals mining"),
    ("gold",            "metals mining"),
    ("silver",          "metals mining"),
    ("platinum",        "metals mining"),
    ("palladium",       "metals mining"),
    ("pgm",             "metals mining"),
    ("niobium",         "metals mining"),
    ("vanadium",        "metals mining"),
    ("uranium",         "metals mining"),
    ("antimony",        "metals mining"),
    ("titanium",        "metals mining"),
    ("gallium",         "metals mining"),
    ("germanium",       "metals mining"),
    ("iron ore",        "metals mining"),
    ("steel",           "metals mining"),
    ("molybdenum",      "metals mining"),
    ("zircon",          "metals mining"),
    ("rutile",          "metals mining"),
    ("graphite",        "industrial minerals"),
    ("silica",          "industrial minerals"),
    ("ball clay",       "industrial minerals"),
    ("gypsum",          "industrial minerals"),
    ("fluorspar",       "industrial minerals"),
    ("potash",          "industrial minerals"),
    ("salt",            "industrial minerals"),
    ("coal",            "coal"),
    ("oil",             "oil and gas"),
    ("gas",             "oil and gas"),
]

# ---------------------------------------------------------------------------
# Commodity abbreviations for product_id generation
# ---------------------------------------------------------------------------
COMMODITY_ABBREV = {
    "lithium carbonate":    "LiCO3",
    "lithium hydroxide":    "LiOH",
    "lithium":              "Li",
    "cobalt metal":         "CoM",
    "cobalt sulfate":       "CoS",
    "cobalt hydroxide":     "CoH",
    "cobalt":               "Co",
    "nickel":               "Ni",
    "copper":               "Cu",
    "zinc":                 "Zn",
    "lead":                 "Pb",
    "tin":                  "Sn",
    "aluminum":             "Al",
    "alumina":              "Al2O3",
    "bauxite":              "Bx",
    "graphite":             "Gr",
    "rare earth":           "REE",
    "ndpr":                 "NdPr",
    "gold":                 "Au",
    "silver":               "Ag",
    "platinum":             "Pt",
    "palladium":            "Pd",
    "uranium":              "U",
    "vanadium":             "V",
    "titanium":             "Ti",
    "iron ore":             "Fe",
    "molybdenum":           "Mo",
    "antimony":             "Sb",
}

# ---------------------------------------------------------------------------
# HTTP config
# ---------------------------------------------------------------------------
REQUEST_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
}
TIMEOUT = 15

# ---------------------------------------------------------------------------
# Canonical name (SOP-DS-002 §7.2)
# ---------------------------------------------------------------------------

def canonical_name(name: str) -> str:
    s = name.strip().lower()
    s = s.replace("&", "and")
    # Remove accents
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    # Remove all punctuation except hyphens and spaces
    s = re.sub(r"[^\w\s-]", "", s)
    # Collapse spaces → hyphens
    s = re.sub(r"\s+", "-", s.strip())
    # Collapse multiple hyphens
    s = re.sub(r"-{2,}", "-", s)
    return s


# ---------------------------------------------------------------------------
# Tier inference
# ---------------------------------------------------------------------------

def infer_tier(type_str: str) -> int:
    """
    Given a free-text Type column (e.g. 'Mine/Refiner/Producer'), return the
    *lowest* (most upstream) tier among all matching tokens.
    """
    if not type_str:
        return 1
    parts = [p.strip().lower() for p in re.split(r"[/,;|]", type_str)]
    best = None
    for part in parts:
        for key, tier in TIER_MAP.items():
            if key in part:
                if best is None or tier < best:
                    best = tier
    return best if best is not None else 1


# ---------------------------------------------------------------------------
# Sector inference
# ---------------------------------------------------------------------------

def infer_sector(products: str, type_str: str = "",
                 sheet_commodity: str = "") -> str:
    """
    Check products → type → sheet context in order, first match wins.
    Falls back to 'metals mining' for this dataset.
    """
    combined = " ".join([products, type_str, sheet_commodity]).lower()
    for keyword, sector in SECTOR_MAP:
        if keyword in combined:
            return sector
    return "metals mining"


# ---------------------------------------------------------------------------
# Product ID generation
# ---------------------------------------------------------------------------

def _commodity_abbrev(product_name: str) -> str:
    pl = product_name.lower()
    for kw, abbrev in COMMODITY_ABBREV.items():
        if kw in pl:
            return abbrev
    # Fallback: uppercase first letters
    words = product_name.strip().split()
    if len(words) == 1:
        return product_name[:4].upper()
    return "".join(w[0].upper() for w in words[:4])


def make_product_id(company_name: str, product_name: str, grade: Optional[str],
                    index: int, country: str = "") -> str:
    """
    Generate a unique product_id matching reference format: LOCATION_COMMODITY[_GRADE]
    e.g.  DRC_KK_COP,  SA_PM_PGM,  USA_Li_01
    Country code (up to 3 chars) + commodity abbreviation + optional grade code.
    """
    # Country → short code
    country_map = {
        "democratic republic of congo": "DRC", "drc": "DRC", "congo": "DRC",
        "south africa": "SA", "usa": "USA", "united states": "USA",
        "australia": "AUS", "canada": "CAN", "china": "CHN",
        "chile": "CHL", "brazil": "BRA", "peru": "PER",
        "indonesia": "IDN", "russia": "RUS", "kazakhstan": "KAZ",
        "mexico": "MEX", "argentina": "ARG", "zambia": "ZMB",
        "angola": "AGO", "ghana": "GHA", "mozambique": "MOZ",
        "tanzania": "TZA", "madagascar": "MDG", "burundi": "BDI",
        "philippines": "PHL", "norway": "NOR", "sweden": "SWE",
        "finland": "FIN", "germany": "DEU", "france": "FRA",
        "belgium": "BEL", "poland": "POL", "switzerland": "CHE",
        "south korea": "KOR", "korea": "KOR", "japan": "JPN",
        "india": "IND", "uk": "GBR", "united kingdom": "GBR",
        "new caledonia": "NCL", "mongolia": "MNG",
    }
    loc = country_map.get(country.lower().strip(), country[:3].upper() if country else "XX")
    abbrev = _commodity_abbrev(product_name)
    grade_code = ""
    if grade:
        grade_code = "_" + "".join(w[0].upper() for w in grade.split()[:2])
    return f"{loc}_{abbrev}{grade_code}_{index:02d}"


# ---------------------------------------------------------------------------
# URL utilities
# ---------------------------------------------------------------------------

def clean_url(url: Optional[str]) -> Optional[str]:
    if not url:
        return None
    url = str(url).strip()
    if not url.startswith("http"):
        url = "https://" + url
    try:
        parsed = urlparse(url)
        return parsed.geturl() if parsed.netloc else None
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Site type inference
# ---------------------------------------------------------------------------

def infer_site_type(text: str, site_name: str = "") -> str:
    combined = (text + " " + site_name).lower()
    if any(k in combined for k in ["recycling", "recycl facility", "recycler"]):
        return "recycling facility"
    if any(k in combined for k in ["peat working", "peat bog"]):
        return "peat workings"
    if any(k in combined for k in ["brine well", "brine"]):
        return "brine well"
    if any(k in combined for k in ["oil well", "onshore oil"]):
        return "oil well"
    if any(k in combined for k in ["gas well", "onshore gas"]):
        return "gas well"
    if any(k in combined for k in ["wharf", "import terminal", "maritime"]):
        return "wharf"
    if any(k in combined for k in ["smelter", "smelt"]):
        return "smelter"
    if any(k in combined for k in ["refinery", "refiner", "refining"]):
        return "refinery"
    if any(k in combined for k in ["processing plant", "lime works", "tilework"]):
        return "processing plant"
    if any(k in combined for k in ["pit", " pit"]):
        return "pit"
    if any(k in combined for k in ["quarry", "quarries"]):
        return "quarry"
    if any(k in combined for k in ["exploration", "explore"]):
        return "exploration site"
    if any(k in combined for k in ["development", "develop", "project"]):
        return "development project"
    if any(k in combined for k in ["mine", "mining"]):
        return "mine"
    return "mine"


# ---------------------------------------------------------------------------
# Manufacturing sites parser
# ---------------------------------------------------------------------------

# Country name patterns for parenthetical extraction
COUNTRY_NAMES = (
    "USA|United States|UK|United Kingdom|Canada|Australia|Germany|France|China|"
    "Japan|Norway|Belgium|Sweden|Finland|Switzerland|South Korea|Korea|Chile|"
    "Brazil|Peru|DRC|Congo|Indonesia|India|Russia|Kazakhstan|Mexico|Argentina|"
    "Mozambique|Tanzania|Madagascar|Burundi|Philippines|New Caledonia|Mongolia|"
    "Poland|Cuba|Bolivia|Serbia|Zambia|Zimbabwe|Ghana|South Africa|Greenland|"
    "Denmark|Spain|Portugal|Netherlands|Ireland|Austria"
)

SITE_KEYWORDS = (
    r"(?:mine|quarry|smelter|refinery|refiner|plant|hub|facility|project|"
    r"operation|complex|deposit|mill|works|farm|processing)"
)


def parse_manufacturing_sites(capacity_raw: str, country: str,
                               company_name: str) -> list[dict]:
    """
    Extract manufacturing_sites from the Capacity/Production Data column.
    Heuristics:
      1. Named sites with explicit country: "Site Name (Country)"
      2. Named sites with site-type keyword: "Escondida mine", "Kokkola refinery"
      3. Mine/project names with embedded country context
    """
    if not capacity_raw:
        if country:
            return [{
                "location": company_name,
                "country": country,
                "site_type": "mine",
                "raw": None,
            }]
        return []

    sites = []
    seen_names = set()

    # Pattern 1: "Site Name (Country)"
    p1 = re.compile(
        r"([A-Z][A-Za-z\s\-']+?)\s+\((" + COUNTRY_NAMES + r")\)",
        re.IGNORECASE
    )
    for m in p1.finditer(capacity_raw):
        site_name = m.group(1).strip()
        site_country = m.group(2).strip()
        if _is_valid_site_name(site_name) and site_name not in seen_names:
            seen_names.add(site_name)
            sites.append({
                "location": site_name,
                "country": site_country,
                "site_type": infer_site_type(capacity_raw, site_name),
                "raw": capacity_raw,
            })

    # Pattern 2: "Name [site-keyword]" where name is properly capitalised
    p2 = re.compile(
        r"([A-Z][A-Za-z\s\-']{2,40}?)\s+" + SITE_KEYWORDS,
        re.IGNORECASE
    )
    for m in p2.finditer(capacity_raw):
        site_name = m.group(1).strip()
        if _is_valid_site_name(site_name) and site_name not in seen_names:
            seen_names.add(site_name)
            # Try to get country from surrounding context
            ctx_start = max(0, m.start() - 60)
            ctx_end = min(len(capacity_raw), m.end() + 60)
            ctx = capacity_raw[ctx_start:ctx_end]
            site_country = _extract_country_from_context(ctx) or country or "Unknown"
            sites.append({
                "location": site_name,
                "country": site_country,
                "site_type": infer_site_type(capacity_raw, site_name),
                "raw": capacity_raw,
            })

    # Pattern 3: Explicit location+country phrase without site keyword
    # e.g. "Salar de Atacama operations"
    p3 = re.compile(
        r"([A-Z][A-Za-z\s\-']{4,50}?)\s+operations?\b",
        re.IGNORECASE
    )
    for m in p3.finditer(capacity_raw):
        site_name = m.group(1).strip()
        if _is_valid_site_name(site_name) and site_name not in seen_names:
            seen_names.add(site_name)
            ctx_start = max(0, m.start() - 60)
            ctx_end = min(len(capacity_raw), m.end() + 60)
            ctx = capacity_raw[ctx_start:ctx_end]
            site_country = _extract_country_from_context(ctx) or country or "Unknown"
            sites.append({
                "location": site_name,
                "country": site_country,
                "site_type": infer_site_type(capacity_raw, site_name),
                "raw": capacity_raw,
            })

    # Fallback: at least one site record using company country
    if not sites:
        sites.append({
            "location": company_name,
            "country": country or "Unknown",
            "site_type": infer_site_type(capacity_raw),
            "raw": capacity_raw,
        })

    return sites


def _is_valid_site_name(name: str) -> bool:
    """Filter out false positives from regex matches."""
    if len(name) < 3 or len(name) > 60:
        return False
    stop_words = {
        "the", "and", "or", "for", "of", "in", "at", "on", "by", "an", "a",
        "world", "global", "major", "multiple", "various", "largest", "first",
        "phase", "tier", "annual", "stage", "capacity", "production", "detailed",
        "only", "china", "usa", "uk", "australia", "canada", "germany", "france",
        "norway", "belgium", "sweden", "finland", "switzerland",
        "expansion", "expanding", "integrated", "commercial",
    }
    if name.lower().strip() in stop_words:
        return False
    # Skip pure numbers
    if re.match(r"^[\d,.\s]+$", name):
        return False
    return True


def _extract_country_from_context(text: str) -> Optional[str]:
    country_pattern = re.compile(
        r"(?<!\w-)\b(" + COUNTRY_NAMES + r")\b", re.IGNORECASE
    )
    # Exclude matches that are preceded by "non-" or "ex-" (negation context)
    for m in country_pattern.finditer(text):
        # Check what precedes the match
        before = text[max(0, m.start()-5):m.start()]
        if re.search(r"non-|ex-", before, re.IGNORECASE):
            continue
        return m.group(1)
    return None


# ---------------------------------------------------------------------------
# Products parser
# ---------------------------------------------------------------------------

GRADE_KEYWORDS = [
    "battery-grade", "battery grade", "LME Grade A", "99.7%", "high-purity",
    "acid grade", "ceramic grade", "metallurgical grade", "industrial grade",
    "food grade", "technical grade", "nuclear grade", "semiconductor grade",
    "spodumene", "hydroxide", "carbonate", "sulfate", "oxide", "concentrate",
]


def parse_products(products_raw: str, company_name: str,
                   category: str, country: str = "") -> list[dict]:
    """
    Build products_offered list — one entry per distinct product name+grade.
    """
    if not products_raw:
        return []

    results = []
    seen = set()
    # Split on commas that are NOT inside parentheses
    # e.g. "Rare earth oxides (NdPr, Dy, Tb), concentrate" → 2 items
    items = re.split(r",\s*(?![^(]*\))", products_raw)

    for idx, item in enumerate(items):
        item = item.strip()
        if not item:
            continue

        # Extract grade hints from item text
        grade = None
        clean_item = item
        for gkw in GRADE_KEYWORDS:
            if gkw.lower() in item.lower():
                grade = gkw
                clean_item = re.sub(re.escape(gkw), "", clean_item,
                                    flags=re.IGNORECASE).strip(" -,")
                break

        product_name = clean_item.strip(" -,").strip()
        if not product_name or len(product_name) < 2:
            continue

        key = (product_name.lower(), (grade or "").lower())
        if key in seen:
            continue
        seen.add(key)

        pid = make_product_id(company_name, product_name, grade, idx + 1, country)

        results.append({
            "product_name": product_name,
            "grade": grade,
            "product_id": pid,
            "category": category.upper() if category else "METALS",
            "source_url": None,   # set below after products list is built
            "datasheet_url": None,
            "cross_graph_material_id": None,
        })

    return results


# ---------------------------------------------------------------------------
# Spreadsheet reader
# ---------------------------------------------------------------------------

def find_header_row(ws, target_cols: tuple = ("Company Name", "Name")) -> Optional[int]:
    for i, row in enumerate(ws.iter_rows(max_row=25, values_only=True), 1):
        cells = [str(c).strip() if c else "" for c in row]
        if any(col in cells for col in target_cols):
            return i
    return None


def col_index(headers: list, *names: str) -> Optional[int]:
    for name in names:
        for i, h in enumerate(headers):
            if h and name.lower() in str(h).lower():
                return i
    return None


def _cell(row: tuple, idx: Optional[int]) -> Optional[str]:
    if idx is None or idx >= len(row):
        return None
    val = row[idx]
    return str(val).strip() if val not in (None, "") else None


def parse_suppliers_from_sheet(ws, sheet_name: str) -> list[dict]:
    """
    Return a list of raw supplier dicts from one worksheet.
    Handles the quirky 'Other mines- recyclers' sheet layout and the
    standard layout used by all other sheets.
    """
    header_row_idx = find_header_row(ws)
    if header_row_idx is None:
        log.debug("No header found in sheet: %s", sheet_name)
        return []

    all_rows = list(ws.iter_rows(min_row=header_row_idx, values_only=True))
    if not all_rows:
        return []

    headers = [str(c).strip() if c else "" for c in all_rows[0]]

    ci = {
        "name":      col_index(headers, "Company Name", "Name"),
        "country":   col_index(headers, "Country", "Headquarters Location"),
        "type":      col_index(headers, "Type"),
        "products":  col_index(headers, "Products"),
        "website":   col_index(headers, "Website"),
        "capacity":  col_index(headers, "Capacity", "Production Data"),
        "transp":    col_index(headers, "Transparency", "Transp", "Data Transparency"),
        "desc":      col_index(headers, "Description"),
        "funding":   col_index(headers, "Funding", "Equity"),
        "done_by":   col_index(headers, "Done by", "Done By"),
        "comments":  col_index(headers, "Comments", "comment"),
    }

    suppliers = []
    for row in all_rows[1:]:
        name_val = _cell(row, ci["name"])
        if not name_val:
            continue
        # Skip URLs, section headings, short strings
        if name_val.startswith("http") or len(name_val) < 2:
            continue
        if name_val.lower() in ("company name", "name", "company",
                                "controlled thermal resources"):
            # Skip repeated header ghost rows (Lithium sheet quirk)
            continue
        # Skip comment/note rows
        if name_val.startswith("//") or name_val.startswith("-->"):
            continue

        country_val = _cell(row, ci["country"])
        # If country column holds a full HQ address ("Carson City, NV"),
        # try to extract just the country portion
        if country_val:
            country_val = _normalise_country(country_val)

        entry: dict = {
            "company_name": name_val,
            "country":      country_val,
            "type":         _cell(row, ci["type"]),
            "products":     _cell(row, ci["products"]),
            "website":      clean_url(_cell(row, ci["website"])),
            "capacity":     _cell(row, ci["capacity"]),
            "transparency": _cell(row, ci["transp"]),
            "description":  _cell(row, ci["desc"]),
            "sheet":        sheet_name,
            # extras fields
            "funding_usd":  _cell(row, ci["funding"]),
            "done_by":      _cell(row, ci["done_by"]),
            "comments":     _cell(row, ci["comments"]),
        }
        suppliers.append(entry)

    return suppliers


def _normalise_country(raw: str) -> str:
    """
    Turn 'Carson City, NV' → 'USA', 'Argentina/ Australia' → 'Argentina'.
    Keep clean country names as-is.
    """
    # US state abbreviations
    us_states = re.compile(
        r"\b(AL|AK|AZ|AR|CA|CO|CT|DE|FL|GA|HI|ID|IL|IN|IA|KS|KY|LA|ME|"
        r"MD|MA|MI|MN|MS|MO|MT|NE|NV|NH|NJ|NM|NY|NC|ND|OH|OK|OR|PA|RI|"
        r"SC|SD|TN|TX|UT|VT|VA|WA|WV|WI|WY|DC)\b"
    )
    if us_states.search(raw):
        return "USA"
    # Take first country in multi-country strings
    first = re.split(r"[/,;]", raw)[0].strip()
    return first


def read_xlsx(path: str) -> list[dict]:
    """
    Load all SUPPLIER_SHEETS and return deduplicated supplier rows.
    Deduplication key = canonical_name; later sheets fill missing fields.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    all_suppliers: dict[str, dict] = {}  # canonical_name → row

    for sheet_name in wb.sheetnames:
        if sheet_name not in SUPPLIER_SHEETS:
            continue
        ws = wb[sheet_name]
        rows = parse_suppliers_from_sheet(ws, sheet_name)
        log.info("Sheet %-35s → %d rows", f"'{sheet_name}'", len(rows))

        for row in rows:
            key = canonical_name(row["company_name"])
            if key not in all_suppliers:
                all_suppliers[key] = row
            else:
                # Merge: fill missing fields from subsequent appearances
                existing = all_suppliers[key]
                for field in ["country", "type", "products", "website",
                              "capacity", "transparency", "description",
                              "funding_usd", "comments"]:
                    if not existing.get(field) and row.get(field):
                        existing[field] = row[field]
                # Accumulate sheet list
                sheets = existing.get("_sheets", [existing.get("sheet", "")])
                if row["sheet"] not in sheets:
                    sheets.append(row["sheet"])
                existing["_sheets"] = sheets

    wb.close()
    result = list(all_suppliers.values())
    log.info("Total unique suppliers after deduplication: %d", len(result))
    return result


# ---------------------------------------------------------------------------
# Website scraping helpers
# ---------------------------------------------------------------------------

# Sub-pages to probe after the homepage
_PROBE_PATHS = [
    "/operations", "/our-operations", "/what-we-do",
    "/about", "/who-we-are", "/about-us",
    "/products", "/our-products",
    "/newsroom", "/news", "/press-releases",
    "/investor-relations", "/investors",
    "/sustainability", "/esg",
    "/locations", "/global-locations",
    "/en/about", "/en/operations",
]


def fetch_page(url: str, session: requests.Session) -> Optional[BeautifulSoup]:
    """Fetch a single URL and return BeautifulSoup, or None on failure."""
    for attempt in range(2):
        try:
            r = session.get(url, timeout=TIMEOUT, headers=REQUEST_HEADERS,
                            allow_redirects=True)
            if r.status_code == 200:
                return BeautifulSoup(r.text, "lxml")
            log.debug("HTTP %s for %s", r.status_code, url)
            return None
        except Exception as e:
            if attempt == 0:
                time.sleep(1)
            else:
                log.debug("Fetch error %s: %s", url, e)
    return None


def fetch_company_pages(url: str, session: requests.Session,
                        max_extra_pages: int = 5) -> tuple[Optional[BeautifulSoup], str]:
    """
    Fetch homepage + up to max_extra_pages sub-pages.
    Returns (merged_soup, fetched_url) — merged_soup combines text from all pages
    by appending their body text into the homepage soup.
    If the homepage itself has almost no text (JS-rendered), tries sub-pages as
    primary source.
    """
    base = url.rstrip("/")
    home_soup = fetch_page(base, session)

    # Measure useful text in homepage
    home_text_len = len(home_soup.get_text().replace(" ", "")) if home_soup else 0
    js_rendered = home_text_len < 500

    if js_rendered:
        log.debug("Homepage appears JS-rendered (%d chars), trying sub-pages", home_text_len)

    # Extract JSON-LD from homepage first (always in raw HTML even on JS sites)
    if home_soup:
        _extract_json_ld_into_soup(home_soup)

    extra_fetched = 0
    for path in _PROBE_PATHS:
        if extra_fetched >= max_extra_pages:
            break
        sub_soup = fetch_page(base + path, session)
        if sub_soup is None:
            continue
        sub_text = sub_soup.get_text().replace(" ", "")
        if len(sub_text) < 200:
            continue
        # Merge: append sub-page body into home_soup
        if home_soup is None:
            home_soup = sub_soup
        else:
            if home_soup.body and sub_soup.body:
                home_soup.body.append(sub_soup.body)
        extra_fetched += 1
        time.sleep(0.3)

    if home_soup is None:
        log.warning("All pages failed for %s", url)
        return None, url

    log.debug("Fetched homepage + %d sub-pages for %s", extra_fetched, base)
    return home_soup, base


def _extract_json_ld_into_soup(soup: BeautifulSoup) -> None:
    """
    Parse JSON-LD <script> blocks and inject their text content as
    a readable <div> into the soup body — makes it visible to all
    downstream text scrapers without any extra code.
    """
    if not soup.body:
        return
    for script in soup.find_all("script", attrs={"type": "application/ld+json"}):
        try:
            data = json.loads(script.string or "")
            flat = _flatten_json_ld(data)
            if flat:
                div = soup.new_tag("div", attrs={"class": "-json-ld"})
                div.string = flat
                soup.body.append(div)
        except Exception:
            pass


def _flatten_json_ld(obj, depth: int = 0) -> str:
    """Recursively flatten a JSON-LD object into a readable string."""
    if depth > 4:
        return ""
    if isinstance(obj, str):
        return obj if len(obj) > 10 else ""
    if isinstance(obj, (int, float)):
        return str(obj)
    if isinstance(obj, list):
        return " ".join(_flatten_json_ld(i, depth + 1) for i in obj if i)
    if isinstance(obj, dict):
        READABLE_KEYS = {
            "name", "description", "headline", "text", "articleBody",
            "addressLocality", "addressCountry", "addressRegion", "telephone",
            "url", "email", "foundingDate", "legalName", "alternateName",
            "slogan", "location", "address", "numberOfEmployees",
        }
        parts = []
        for k, v in obj.items():
            if k.startswith("@"):
                continue
            key_lower = k.lower()
            if key_lower in READABLE_KEYS or any(rk in key_lower for rk in READABLE_KEYS):
                t = _flatten_json_ld(v, depth + 1)
                if t:
                    parts.append(f"{k}: {t}")
            else:
                t = _flatten_json_ld(v, depth + 1)
                if t:
                    parts.append(t)
        return " | ".join(parts)
    return ""


def _clean_text(el) -> str:
    return re.sub(r"\s+", " ", el.get_text(separator=" ")).strip()


def scrape_description(soup: BeautifulSoup, company_name: str) -> Optional[str]:
    """Extract company description; prefer meta description, then structured text."""
    # 1. Meta description tag
    meta = soup.find("meta", attrs={"name": "description"})
    if meta and meta.get("content"):
        desc = meta["content"].strip()
        if len(desc) > 40:
            return desc

    # 2. og:description
    og = soup.find("meta", property="og:description")
    if og and og.get("content"):
        desc = og["content"].strip()
        if len(desc) > 40:
            return desc

    # 3. First substantial paragraph mentioning the company
    first_word = company_name.lower().split()[0]
    for p in soup.find_all("p"):
        text = _clean_text(p)
        if len(text) > 80 and first_word in text.lower():
            return text[:500]

    # 4. First paragraph of sufficient length anywhere
    for p in soup.find_all("p"):
        text = _clean_text(p)
        if len(text) > 100:
            return text[:500]

    return None


def scrape_hq(soup: BeautifulSoup) -> Optional[str]:
    """Try to extract headquarters from footer/contact sections."""
    full_text = soup.get_text(separator=" ")
    patterns = [
        # "Headquartered in City, Country"
        r"headquarter(?:ed|s)?\s+in\s+([A-Z][a-zA-Z\s]+,\s*[A-Z][a-zA-Z\s]+)",
        # "Head office: City, Country"
        r"head\s+office[:\s]+([A-Z][a-zA-Z\s]+,\s*[A-Z][a-zA-Z\s]+)",
        # Address block pattern: "City, Country"
        (r"([A-Z][a-zA-Z]+(?: [A-Z][a-zA-Z]+)?,\s*(?:"
         + COUNTRY_NAMES.replace("|", "|") + r"))"),
    ]
    for pat in patterns:
        m = re.search(pat, full_text, re.IGNORECASE)
        if m:
            return m.group(1).strip()[:80]
    return None


def scrape_links(soup: BeautifulSoup, base_url: str,
                 keywords: list[str]) -> Optional[str]:
    """Find a link whose href or text matches any of the keywords."""
    for a in soup.find_all("a", href=True):
        href = a["href"].lower()
        text = a.get_text().lower()
        if any(kw in href or kw in text for kw in keywords):
            full = urljoin(base_url, a["href"])
            if full.startswith("http"):
                return full
    return None


def scrape_ticker(soup: BeautifulSoup) -> tuple[Optional[str], Optional[str]]:
    """Extract ticker symbol and exchange."""
    text = soup.get_text(separator=" ")
    patterns = [
        r"\b(NYSE|NASDAQ|ASX|LSE|TSX|HKEX|HKEx|OTCQX|OTC|JSE|NZX)\s*:?\s*([A-Z]{1,6})\b",
        r"\b([A-Z]{1,6})\s*[:|]\s*(NYSE|NASDAQ|ASX|LSE|TSX|HKEX|HKEx|OTCQX|OTC)\b",
    ]
    for pat in patterns:
        m = re.search(pat, text)
        if m:
            g = m.groups()
            if g[0] in ("NYSE", "NASDAQ", "ASX", "LSE", "TSX", "HKEX",
                        "HKEx", "OTCQX", "OTC", "JSE", "NZX"):
                return g[1], g[0]   # ticker, exchange
            return g[0], g[1]
    return None, None


def scrape_certifications(soup: BeautifulSoup) -> list[str]:
    """Look for known certification names in page text."""
    text = soup.get_text(separator=" ")
    known = [
        "Copper Mark", "IRMA", "ISO 14001", "ISO 9001", "ISO 45001",
        "LME Responsible Sourcing", "LBMA", "RMI", "GRI", "TCFD",
        "Responsible Minerals Initiative", "RMAP",
        "Initiative for Responsible Mining Assurance",
    ]
    return [cert for cert in known if cert.lower() in text.lower()]


def scrape_annual_report(soup: BeautifulSoup, base_url: str) -> Optional[str]:
    return scrape_links(soup, base_url,
                        ["annual report", "annual-report", "annual_report"])


def scrape_investor_url(soup: BeautifulSoup, base_url: str) -> Optional[str]:
    return scrape_links(soup, base_url, ["investor", "/ir/", "/ir "])


def scrape_sustainability_url(soup: BeautifulSoup,
                              base_url: str) -> Optional[str]:
    return scrape_links(soup, base_url,
                        ["sustainability", "esg", "environment", "responsibility",
                         "climate", "responsible"])



# ---------------------------------------------------------------------------
# Claude LLM extraction (enhanced mining sites + products)
# ---------------------------------------------------------------------------

LLM_SYSTEM = """You are a critical materials supply chain data extraction specialist.
You extract structured supplier intelligence from company website text.
You ALWAYS follow the BGS Supplier Graph Schema structure exactly.
You output ONLY valid JSON — no preamble, no markdown fences.

FIELD RULES:
- manufacturing_sites: array of objects {location, country, site_type, raw}
  site_type must be ONE OF: mine|quarry|pit|refinery|smelter|processing plant|
  handling site|wharf|recycling facility|exploration site|laboratory|development project
- products_offered: array of objects {product_name, grade, product_id, category,
  source_url, datasheet_url, cross_graph_material_id}
  Separate battery-grade vs technical-grade vs concentrate as DIFFERENT entries.
  product_id format: {COUNTRY_CODE}_{COMMODITY_ABBREV}_{INDEX:02d}
- annual_production: array of {commodity, volume, unit, year, notes}
- jv_stakes: array of {site_name, ownership_pct, jv_partners, country, commodity}

NEVER invent data. If not on the website, use null or [].
NEVER use plain strings in arrays — always use proper objects."""


def claude_extract(page_text: str, company_name: str, category: str,
                   website_url: str, api_key: str) -> dict:
    """
    Call Claude to extract structured manufacturing_sites, products_offered,
    and extras from scraped website text.
    Returns a partial record dict with only the Claude-extracted fields.
    """
    if not _HAS_ANTHROPIC:
        raise RuntimeError("pip install anthropic to use --use-llm")

    client = anthropic.Anthropic(api_key=api_key)
    prompt = f"""Company: {company_name}
Category: {category}
Website: {website_url}

WEBSITE CONTENT (from multiple pages):
---
{page_text[:50000]}
---

Extract and return a JSON object with ONLY these keys:
{{
  "company_description": "...",
  "headquarters_location": "City, Country",
  "manufacturing_sites": [
    {{"location":"...", "country":"...", "site_type":"mine", "raw":"verbatim text"}}
  ],
  "products_offered": [
    {{"product_name":"...", "grade":"...", "product_id":"XX_ABBREV_01",
      "category":"{category.upper()}", "source_url":"{website_url}",
      "datasheet_url":null, "cross_graph_material_id":null}}
  ],
  "annual_production": [
    {{"commodity":"...", "volume":"...", "unit":"tonnes", "year":"2024", "notes":"..."}}
  ],
  "jv_stakes": [
    {{"site_name":"...", "ownership_pct":44.0, "jv_partners":[], "country":"...", "commodity":"..."}}
  ],
  "certifications": [],
  "ticker_symbol": null,
  "stock_exchange": null,
  "sustainability_report_url": null,
  "annual_report_url": null,
  "investor_relations_url": null
}}"""

    response = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=3000,
        system=LLM_SYSTEM,
        messages=[{"role": "user", "content": prompt}],
    )
    raw_text = response.content[0].text if response.content else ""
    clean = raw_text.strip()
    if clean.startswith("```"):
        clean = re.sub(r"^```(?:json)?\s*", "", clean)
        clean = re.sub(r"\s*```$", "", clean)
    return json.loads(clean.strip())



# ---------------------------------------------------------------------------
# Record builder
# ---------------------------------------------------------------------------

def build_record(raw: dict, soup: Optional[BeautifulSoup],
                 fetched_url: Optional[str]) -> dict:
    """
    Combine spreadsheet row + scraped website data into a BGS schema record.
    Anything that does not fit BGS schema fields goes into `extras`.
    """
    name        = raw["company_name"]
    country     = (raw.get("country") or "").strip()
    type_str    = raw.get("type") or ""
    products_raw = raw.get("products") or ""
    capacity_raw = raw.get("capacity") or ""
    transparency = raw.get("transparency")
    desc_xl     = raw.get("description")
    sheet       = raw.get("sheet", "")
    sheets      = raw.get("_sheets", [sheet])

    sheet_commodity = SHEET_COMMODITY_CONTEXT.get(sheet, "")
    category = sheet_commodity.upper() if sheet_commodity else "METALS"

    # ---- BGS Schema fields ----
    record: dict = {
        "supplier_id":            None,
        "duns_number":            None,
        "company_name":           name,
        "canonical_name":         canonical_name(name),
        "headquarters_location":  country if country else None,
        "website":                fetched_url or raw.get("website"),
        "company_description":    desc_xl,
        "industry_sector":        infer_sector(products_raw, type_str,
                                               sheet_commodity),
        "supply_chain_tier":      infer_tier(type_str),
        "typical_lead_time_days": None,
        "is_verified":            False,
        "manufacturing_sites":    parse_manufacturing_sites(
                                      capacity_raw, country, name),
        "products_offered":       parse_products(products_raw, name, category, country),
        "certification_references": [],
        "certifications_raw":     None,
        "regulation_references":  [],
        "data_completeness_flags": DATA_COMPLETENESS_FLAGS.copy(),
        "sources":                [],   # populated below once website URL is resolved
        # extras: list of dicts — one dict per logical group of extra fields
        "extras": [{}],
    }

    # ---- Enrich from website ----
    if soup and fetched_url:
        # Description
        if not record["company_description"]:
            record["company_description"] = scrape_description(soup, name)

        # HQ location (override spreadsheet only if scrape is more specific)
        scraped_hq = scrape_hq(soup)
        if scraped_hq and "," in scraped_hq:
            record["headquarters_location"] = scraped_hq
        elif not record["headquarters_location"] and country:
            record["headquarters_location"] = country

        # Ticker
        ticker, exchange = scrape_ticker(soup)
        if ticker:
            record["extras"][0]["ticker_symbol"] = ticker
            record["extras"][0]["stock_exchange"] = exchange

        # Important link URLs
        ir_url = scrape_investor_url(soup, fetched_url)
        if ir_url:
            record["extras"][0]["investor_relations_url"] = ir_url

        sus_url = scrape_sustainability_url(soup, fetched_url)
        if sus_url:
            record["extras"][0]["sustainability_report_url"] = sus_url

        ar_url = scrape_annual_report(soup, fetched_url)
        if ar_url:
            record["extras"][0]["annual_report_url"] = ar_url

        # Certifications
        certs = scrape_certifications(soup)
        if certs:
            record["extras"][0]["certifications_found"] = certs
            record["certifications_raw"] = "; ".join(certs)

    # ---- Backfill source_url on products_offered with the company website ----
    company_website = fetched_url or raw.get("website")
    for p in record["products_offered"]:
        if p["source_url"] is None and company_website:
            p["source_url"] = company_website

    # ---- Build sources from actual company URL ----
    company_url = fetched_url or raw.get("website")
    if soup and company_url:
        # Use the scraped page <title> as source_name, fall back to company name
        title_tag = soup.find("title")
        page_title = title_tag.get_text().strip() if title_tag else name
        if not page_title or len(page_title) < 3:
            page_title = name
        record["sources"].append(make_source(page_title, company_url, "tier1"))
    elif company_url:
        # No soup (fetch failed) — still record the URL with company name
        record["sources"].append(make_source(name, company_url, "tier1"))
    else:
        # No URL at all — mark as spreadsheet-only
        record["sources"].append(make_source(
            " Internal Critical Materials Suppliers Spreadsheet",
            None, "tier2"
        ))

    # ---- Extras from spreadsheet ----
    if capacity_raw:
        record["extras"][0]["capacity_production_data_raw"] = capacity_raw

    if transparency:
        record["extras"][0]["data_transparency_level"] = transparency

    if products_raw:
        record["extras"][0]["primary_commodities_raw"] = products_raw

    if type_str:
        record["extras"][0]["company_type_raw"] = type_str

    if raw.get("funding_usd"):
        try:
            record["extras"][0]["total_equity_funding_usd"] = float(raw["funding_usd"])
        except (ValueError, TypeError):
            record["extras"][0]["total_equity_funding_usd"] = raw["funding_usd"]

    if raw.get("comments"):
        record["extras"][0]["analyst_comments"] = raw["comments"]

    if raw.get("done_by"):
        record["extras"][0]["sourced_by"] = raw["done_by"]

    # Track which sheets contributed to this record
    record["extras"][0]["source_sheets"] = sheets if len(sheets) > 1 else sheets[0]

    return record


# ---------------------------------------------------------------------------
# Worker
# ---------------------------------------------------------------------------

def process_one(raw: dict, session: requests.Session,
                delay: float, skip_fetch: bool,
                use_llm: bool = False, api_key: str = "") -> dict:
    url = raw.get("website")
    soup = None
    fetched_url = url

    if url and not skip_fetch:
        time.sleep(delay)
        # Use multi-page fetcher: homepage + sub-pages + JSON-LD
        soup, fetched_url = fetch_company_pages(url, session)
        if soup is None:
            log.warning("Could not fetch %s — building from spreadsheet only", url)

    record = build_record(raw, soup, fetched_url)

    # LLM enrichment: override heuristic sites+products with Claude extraction
    if use_llm and api_key and soup and fetched_url:
        try:
            page_text = soup.get_text(separator="\n")
            page_text = re.sub(r"\n{3,}", "\n\n", page_text).strip()
            sheet = raw.get("sheet", "")
            category = SHEET_COMMODITY_CONTEXT.get(sheet, "metals mining")
            llm_data = claude_extract(page_text, raw["company_name"],
                                      category, fetched_url, api_key)
            # Merge LLM data into record (LLM wins over heuristics for these fields)
            if llm_data.get("manufacturing_sites"):
                record["manufacturing_sites"] = llm_data["manufacturing_sites"]
            if llm_data.get("products_offered"):
                record["products_offered"] = llm_data["products_offered"]
            if llm_data.get("company_description") and not record.get("company_description"):
                record["company_description"] = llm_data["company_description"]
            if llm_data.get("headquarters_location") and "," in str(llm_data.get("headquarters_location","")):
                record["headquarters_location"] = llm_data["headquarters_location"]
            # Extras from LLM
            ex = record["extras"][0]
            for field in ["annual_production","jv_stakes","certifications",
                          "ticker_symbol","stock_exchange","sustainability_report_url",
                          "annual_report_url","investor_relations_url"]:
                if llm_data.get(field):
                    ex[field] = llm_data[field]
            log.info("  LLM %-43s  %d sites  %d products",
                     raw["company_name"][:43],
                     len(record["manufacturing_sites"]),
                     len(record["products_offered"]))
        except Exception as e:
            log.warning("LLM extraction failed for %s: %s", raw["company_name"], e)

    log.info(
        "✓ %-45s  tier=%s  sector=%s  sites=%d  products=%d",
        record["company_name"][:45],
        record["supply_chain_tier"],
        record["industry_sector"],
        len(record["manufacturing_sites"]),
        len(record["products_offered"]),
    )
    return record


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Extract critical materials supplier records → BGS schema JSON"
    )
    parser.add_argument(
        "--xlsx", default="Intern-_Critical_Materials_Suppliers.xlsx",
        help="Path to the input xlsx file"
    )
    parser.add_argument(
        "--out", default="suppliers",
        help="Output directory for per-company JSON files (default: ./suppliers/)"
    )
    parser.add_argument(
        "--limit", type=int, default=None,
        help="Only process the first N suppliers (useful for testing)"
    )
    parser.add_argument(
        "--workers", type=int, default=4,
        help="Number of parallel fetch threads (default 4)"
    )
    parser.add_argument(
        "--delay", type=float, default=1.0,
        help="Seconds to wait between requests per worker thread (default 1.0)"
    )
    parser.add_argument(
        "--skip-fetch", action="store_true",
        help="Skip website fetching; build records from spreadsheet data only"
    )
    parser.add_argument(
        "--use-llm", action="store_true",
        help="Use Claude AI to extract manufacturing_sites and products_offered from website text"
    )
    parser.add_argument(
        "--api-key", default="",
        help="Anthropic API key for --use-llm (or set ANTHROPIC_API_KEY env var)"
    )
    parser.add_argument(
        "--project-name", default="",
        help="Project name for the output folder README (optional)"
    )
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        log.error("File not found: %s", xlsx_path)
        return

    log.info("Reading %s …", xlsx_path)
    suppliers = read_xlsx(str(xlsx_path))

    if args.limit:
        suppliers = suppliers[: args.limit]
        log.info("Limiting to first %d suppliers", args.limit)

    # Resolve API key
    import os
    api_key = args.api_key or os.getenv("ANTHROPIC_API_KEY", "")
    if args.use_llm and not api_key:
        log.error("--use-llm requires an Anthropic API key (--api-key or ANTHROPIC_API_KEY env)")
        return
    if args.use_llm and not _HAS_ANTHROPIC:
        log.error("pip install anthropic to use --use-llm")
        return

    records = []
    session = requests.Session()
    session.headers.update(REQUEST_HEADERS)

    with ThreadPoolExecutor(max_workers=args.workers) as pool:
        futures = {
            pool.submit(process_one, raw, session, args.delay,
                        args.skip_fetch, args.use_llm, api_key): raw
            for raw in suppliers
        }
        for fut in as_completed(futures):
            try:
                records.append(fut.result())
            except Exception as e:
                raw = futures[fut]
                log.error("Failed for %s: %s", raw.get("company_name"), e)

    # Sort deterministically
    records.sort(key=lambda r: r["canonical_name"])

    # ── Project folder output ──────────────────────────────────────────────
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    project_name = args.project_name or "critical_materials_intelligence"
    project_slug = re.sub(r"[^a-z0-9_]", "_", project_name.lower().strip())
    out_dir = Path(args.out) / f"{project_slug}_{timestamp}"
    records_dir = out_dir / "records"
    records_dir.mkdir(parents=True, exist_ok=True)

    # One JSON per supplier
    for record in records:
        filepath = records_dir / f"{record['canonical_name']}.json"
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(record, f, indent=2, ensure_ascii=False)

    # Combined JSON (all records in one array — upload this to )
    combined_path = out_dir / "combined_suppliers.json"
    with open(combined_path, "w", encoding="utf-8") as f:
        json.dump(records, f, indent=2, ensure_ascii=False)

    # README.md
    readme_path = out_dir / "README.md"
    with open(readme_path, "w", encoding="utf-8") as f:
        f.write(f"""# {project_name.title()} — Extraction Output

Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
Source: {args.xlsx}
Method: {"Excel + Website Scraping + Claude LLM" if args.use_llm else "Excel + Website Scraping" if not args.skip_fetch else "Excel only (--skip-fetch)"}

## Summary
- Total suppliers extracted: {len(records)}
- Records folder: `records/` ({len(records)} individual JSON files)
- Combined file: `combined_suppliers.json` (upload this to )

## Files
| File | Contents |
|------|----------|
| `combined_suppliers.json` | All {len(records)} records in one array — for  upload |
| `records/*.json` | One file per company — for individual review/sharing |
| `README.md` | This file |

## Upload to 
```bash
python seed_critical_materials.py \
  --combined combined_suppliers.json \
  --url https://-platform-production.up.railway.app \
  --email admin@yourorg.com --password yourpass
```

## Tier Breakdown
""")
        for t, n in sorted(tiers.items()):
            label = {1: "Tier 1 — Extraction/Mining", 2: "Tier 2 — Processing/Refining", 3: "Tier 3 — Trading"}.get(t, f"Tier {t}")
            f.write(f"- {label}: {n}\n")
        f.write("\n## Sectors\n")
        for s, n in sorted(sectors.items(), key=lambda x: -x[1]):
            f.write(f"- {s}: {n}\n")

    # Export report JSON
    report_path = out_dir / "export_report.json"
    with open(report_path, "w", encoding="utf-8") as f:
        json.dump({
            "generated_at": datetime.now().isoformat(),
            "source_file": str(args.xlsx),
            "extraction_method": "llm+web" if args.use_llm else "web" if not args.skip_fetch else "excel_only",
            "total_records": len(records),
            "with_manufacturing_sites": with_sites,
            "with_products_offered": with_products,
            "tiers": tiers,
            "sectors": sectors,
        }, f, indent=2)

    log.info("Project folder → %s/", out_dir)
    log.info("  records/         %d individual JSON files", len(records))
    log.info("  combined_suppliers.json  (upload this to )")
    log.info("  README.md        summary and upload instructions")
    log.info("  export_report.json")

    # ---- Summary ----
    tiers: dict[int, int] = {}
    sectors: dict[str, int] = {}
    with_sites = 0
    with_products = 0
    for r in records:
        t = r["supply_chain_tier"]
        tiers[t] = tiers.get(t, 0) + 1
        s = r["industry_sector"]
        sectors[s] = sectors.get(s, 0) + 1
        if r["manufacturing_sites"]:
            with_sites += 1
        if r["products_offered"]:
            with_products += 1

    print("\n" + "=" * 50)
    print(f"  EXTRACTION SUMMARY")
    print("=" * 50)
    print(f"  Total records       : {len(records)}")
    print(f"  Output directory    : {out_dir}/")
    print(f"  With mfg sites      : {with_sites}")
    print(f"  With products       : {with_products}")
    print(f"  Tier 1 (extraction) : {tiers.get(1, 0)}")
    print(f"  Tier 2 (processing) : {tiers.get(2, 0)}")
    print(f"  Tier 3 (trading)    : {tiers.get(3, 0)}")
    print("\n  Sectors:")
    for s, n in sorted(sectors.items(), key=lambda x: -x[1]):
        print(f"    {s:<35} {n}")
    print("=" * 50)


if __name__ == "__main__":
    main()
