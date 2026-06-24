#!/usr/bin/env python3
"""
seed_critical_materials.py — Seed Xtrium with 141 Critical Materials companies
Creates 1 project, 1 schema, 141 sources (one per company with its website_url
set for LLM verification), and 141 records pre-loaded.

Usage:
  python seed_critical_materials.py \
    --url https://xtrium-platform-production.up.railway.app \
    --email admin@yourorg.com --password yourpass

Requirements: pip install requests openpyxl
"""
import sys, json, re, argparse, io, time
from pathlib import Path

try:
    import requests
except ImportError:
    sys.exit("pip install requests")
try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("pip install openpyxl")

SCHEMA_DEF = {
    "name": "Critical Materials Supplier Schema v1.0",
    "grouping_key": "company_name",
    "extraction_instructions": (
        "ONE record per company. Required: company_name, canonical_name, industry_sector, "
        "supply_chain_tier (1=mine/primary, 2=refiner/smelter/processor/recycler, 3=trader), "
        "is_verified (always false). CANONICAL NAME: lowercase, &→and, spaces→hyphens, "
        "remove punctuation. Everything else is optional."
    ),
    "fields": [
        {"name": "supplier_id", "type": "string", "required": False, "fixed_value": None},
        {"name": "duns_number", "type": "string", "required": False, "fixed_value": None},
        {"name": "company_name", "type": "string", "required": True},
        {"name": "canonical_name", "type": "string", "required": True},
        {"name": "headquarters_location", "type": "string", "required": False},
        {"name": "website", "type": "string", "required": False},
        {"name": "company_description", "type": "string", "required": False},
        {"name": "industry_sector", "type": "string", "required": True,
         "enum": ["metals mining","industrial minerals","construction minerals",
                  "cement and lime","coal","oil and gas","recycled aggregates","peat","mineral handling"]},
        {"name": "supply_chain_tier", "type": "integer", "required": True},
        {"name": "typical_lead_time_days", "type": "string", "required": False, "fixed_value": None},
        {"name": "is_verified", "type": "boolean", "required": True, "fixed_value": False},
        {"name": "type_description", "type": "string", "required": False},
        {"name": "products_raw", "type": "string", "required": False},
        {"name": "data_transparency_level", "type": "string", "required": False,
         "enum": ["High","Medium-High","Medium","Low-Medium","Low"]},
        {"name": "manufacturing_sites", "type": "array", "required": False},
        {"name": "products_offered", "type": "array", "required": False},
        {"name": "certification_references", "type": "array", "required": False, "fixed_value": []},
        {"name": "certifications_raw", "type": "string", "required": False, "fixed_value": None},
        {"name": "regulation_references", "type": "array", "required": False, "fixed_value": []},
        {"name": "data_completeness_flags", "type": "object", "required": False},
        {"name": "sources", "type": "array", "required": False},
    ]
}

CAT_SECTOR = {
    'Lithium Producers':'industrial minerals','Rare Earth Elements':'industrial minerals',
    'Cobalt Producers':'metals mining','Copper Producers':'metals mining',
    'Nickel Producers':'metals mining','Graphite Producers':'industrial minerals',
    'Aluminum Producers':'construction minerals','Zinc Lead Tin Producers':'metals mining',
    'Multi-Metal Refiners':'metals mining','Specialty Critical Materials':'industrial minerals',
    'Other mines- recyclers':'recycled aggregates',
}

def mkcn(name):
    n = str(name).lower().strip().replace('&','and')
    n = re.sub(r'\s+','-',n); n = re.sub(r'[^a-z0-9\-]','',n)
    return re.sub(r'-+','-',n).strip('-')

def tier(t):
    if not t: return 1
    t = t.lower()
    if any(x in t for x in ['mine','explorer','developer','producer']): return 1
    if any(x in t for x in ['refiner','smelter','recycler','processor']): return 2
    return 3 if 'trader' in t else 1

def url(w):
    if not w: return None
    w = str(w).strip().split(',')[0].strip()
    if not w.startswith('http') and '.' in w: w = 'https://'+w
    return w if w.startswith('http') else None

def norm_transp(t):
    if not t: return None
    for lv in ['High','Medium-High','Medium','Low-Medium','Low']:
        if lv.lower() in t.lower(): return lv
    return None

def parse_excel(xlsx):
    wb = load_workbook(xlsx, read_only=True)
    companies = []
    for sheet in wb.sheetnames:
        if sheet == 'Material web indexes': continue
        ws = wb[sheet]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        sector = CAT_SECTOR.get(sheet,'metals mining')
        hidx = next((i for i,r in enumerate(rows) if len(r)>1 and r[1] and 'Country' in str(r[1])),None)

        if sheet == 'Other mines- recyclers':
            for row in rows:
                if not row or not row[0]: continue
                name = str(row[0]).strip()
                if name.startswith('//') or name in ('Name','Company Name'): continue
                c1 = str(row[1]).strip() if len(row)>1 and row[1] else ''
                if c1.startswith('http'):
                    companies.append({'category':sheet,'company_name':name,'website':c1,
                        'headquarters_location':None,'company_description':None,
                        'type_description':'Recycler/Other','products_raw':None,
                        'data_transparency_level':None,'industry_sector':sector,'supply_chain_tier':2})
                elif c1 and len(c1)>20:
                    hq = str(row[2]).strip() if len(row)>2 and row[2] else None
                    companies.append({'category':sheet,'company_name':name,'website':None,
                        'headquarters_location':hq,'company_description':c1,
                        'type_description':'Recycler','products_raw':None,
                        'data_transparency_level':None,'industry_sector':sector,'supply_chain_tier':2})
        elif hidx is not None:
            for row in rows[hidx+1:]:
                if not row or not row[0] or str(row[0]).strip() in ('','Company Name'): continue
                name = str(row[0]).strip()
                type_str = str(row[2]).strip() if len(row)>2 and row[2] else None
                desc = str(row[5]).strip() if len(row)>5 and row[5] and len(str(row[5]))<800 else None
                companies.append({'category':sheet,'company_name':name,
                    'website':url(row[4] if len(row)>4 else None),
                    'headquarters_location':str(row[1]).strip() if len(row)>1 and row[1] else None,
                    'type_description':type_str,
                    'products_raw':str(row[3]).strip() if len(row)>3 and row[3] else None,
                    'company_description':desc,
                    'data_transparency_level':norm_transp(str(row[6]).strip() if len(row)>6 and row[6] else None),
                    'industry_sector':sector,'supply_chain_tier':tier(type_str)})
    return companies

def build_record(c):
    name = c['company_name']; w = c.get('website'); cn = mkcn(name)
    prod = c.get('products_raw') or ''
    return {
        'supplier_id':None,'duns_number':None,'company_name':name,'canonical_name':cn,
        'headquarters_location':c.get('headquarters_location'),'website':w,
        'company_description':c.get('company_description'),
        'industry_sector':c.get('industry_sector','metals mining'),
        'supply_chain_tier':c.get('supply_chain_tier',1),'typical_lead_time_days':None,
        'is_verified':False,'type_description':c.get('type_description'),
        'products_raw':prod,'data_transparency_level':c.get('data_transparency_level'),
        'manufacturing_sites':[],'products_offered':[{
            'product_name':prod,'grade':None,'product_id':cn[:15]+'_01',
            'category':c.get('category',''),'source_url':w,'datasheet_url':None,'cross_graph_material_id':None
        }] if prod else [],
        'certification_references':[],'certifications_raw':None,'regulation_references':[],
        'data_completeness_flags':{'review_score':'manual_only','defect_rate_ppm':'manual_only',
            'on_time_delivery_rate':'manual_only','pricing':'api_only','inventory_levels':'api_only'},
        'sources':[{'source_name':'Xtrium Critical Materials Intelligence — Intern Research 2025',
            'source_url':w or 'https://xtrium.ai','doi':None,'tier':'tier2'}],
    }

class API:
    def __init__(self,base,email,pw):
        self.b=base.rstrip('/'); self.s=requests.Session()
        self.s.headers['Content-Type']='application/json'
        r=self.s.post(f'{self.b}/api/v1/auth/login',json={'email':email,'password':pw})
        r.raise_for_status(); self.s.headers['Authorization']=f"Bearer {r.json()['access_token']}"
        print(f'✓ Authenticated as {email}')
    def get(self,p,**kw): r=self.s.get(f'{self.b}{p}',**kw); r.raise_for_status(); return r.json()
    def post(self,p,**kw): r=self.s.post(f'{self.b}{p}',**kw); r.raise_for_status(); return r.json()
    def upload(self,sid,record,fname):
        content=json.dumps([record],ensure_ascii=False,indent=2).encode()
        r=self.s.post(f'{self.b}/api/v1/sources/{sid}/upload',
            files={'file':(fname,io.BytesIO(content),'application/json')},
            headers={k:v for k,v in self.s.headers.items() if k!='Content-Type'})
        r.raise_for_status(); return r.json()

def run(args):
    print('\n[1/4] Parsing Excel…')
    xlsx = args.xlsx or '/mnt/user-data/uploads/Intern-_Critical_Materials_Suppliers.xlsx'
    if not Path(xlsx).exists(): sys.exit(f'Excel not found: {xlsx}')
    companies = parse_excel(xlsx)
    print(f'  {len(companies)} companies parsed')

    api = API(args.url,args.email,args.password)

    print('\n[2/4] Project + schema…')
    projects = api.get('/api/v1/projects').get('items',[])
    proj = next((p for p in projects if p['name']=='Critical Materials Intelligence'),None)
    if not proj:
        proj = api.post('/api/v1/projects',json={'name':'Critical Materials Intelligence',
            'description':'141 critical materials suppliers. Each source = one company with its website set — use "LLM Verify vs Website" to cross-check each record against the real company site.',
            'status':'active'})
        print(f'  ✓ Created project')
    pid = proj['id']

    schemas = api.get('/api/v1/schemas',params={'project_id':pid})
    schema = next((s for s in schemas if s['name']==SCHEMA_DEF['name']),None)
    if not schema:
        schema = api.post(f'/api/v1/schemas/{pid}',json={'name':SCHEMA_DEF['name'],'definition':SCHEMA_DEF})
        print(f'  ✓ Created schema')
    sid = schema['id']

    print(f'\n[3/4] Creating {len(companies)} sources + uploading records…')
    existing_raw = api.get('/api/v1/sources',params={'project_id':pid})
    existing = {s['name'] for s in (existing_raw if isinstance(existing_raw, list) else existing_raw.get('items',[]))}
    valid=0; invalid=0; created=0; skipped=0

    for i,c in enumerate(companies):
        name = c['company_name']
        if name in existing: skipped+=1; continue
        try:
            src = api.post('/api/v1/sources',params={'project_id':pid},json={
                'name':name,
                'description':f"{c.get('type_description','')} | {c.get('headquarters_location','?')} | Transparency: {c.get('data_transparency_level','?')}",
                'schema_id':sid,
                'website_url':c.get('website'),  # ← LLM verification target
            })
            existing.add(name)
            result = api.upload(src['id'],build_record(c),f"{mkcn(name)}.json")
            valid+=result.get('valid_rows',0); invalid+=result.get('invalid_rows',0); created+=1
            if (i+1)%15==0 or i==len(companies)-1:
                print(f'  [{i+1}/{len(companies)}] {valid} valid, {invalid} need fixes')
            time.sleep(0.1)
        except Exception as e:
            body = ''
            try:
                body = e.response.text[:300] if hasattr(e,'response') and e.response else ''
            except: pass
            print(f'  ✗ {name}: {e} | {body}'); invalid+=1

    print(f'\n[4/4] Done!')
    print(f'  Sources created: {created}  |  Skipped (existing): {skipped}')
    print(f'  Valid records:   {valid}  |  Needs review: {invalid}')
    print(f'\n→ Open Xtrium → Projects → Critical Materials Intelligence')
    print(f'→ Each company has its own source. Click any source → "LLM Verify vs Website"')
    print(f'  to auto-check the extracted data against the company\'s real website.')

def main():
    p=argparse.ArgumentParser(description='Seed Xtrium with 141 Critical Materials Suppliers')
    p.add_argument('--url',default='https://xtrium-platform-production.up.railway.app')
    p.add_argument('--email',required=True); p.add_argument('--password',required=True)
    p.add_argument('--xlsx',help='Path to Excel file')
    main_args=p.parse_args(); run(main_args)

if __name__=='__main__': main()
