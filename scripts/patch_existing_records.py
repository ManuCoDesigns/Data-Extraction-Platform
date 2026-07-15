#!/usr/bin/env python3
"""
patch_existing_records.py — Update already-seeded records with Excel data
=========================================================================
Reads the Critical Materials Excel, finds each company's existing source
and record in , and PATCHes the record with the correctly mapped
field values — WITHOUT creating duplicates.

Run this when:
  - Records are already seeded but fields are missing/wrong
  - You updated the schema and want existing records to reflect it
  - You want to enrich seeded records with fresh Excel data

Usage:
  python scripts/patch_existing_records.py \
    --url https://-platform-production.up.railway.app \
    --email garima@careerflow.ai \
    --password garima@careerflow.ai \
    --xlsx "Intern- Critical_Materials_Suppliers.xlsx"

  # Dry run (see what would change without saving):
  python scripts/patch_existing_records.py ... --dry-run

  # Only patch a specific project:
  python scripts/patch_existing_records.py ... --project "Critical Materials Intelligence"
"""
import sys, json, re, argparse, time, unicodedata
from pathlib import Path

try:
    import requests
    from openpyxl import load_workbook
except ImportError:
    sys.exit("pip install requests openpyxl")


CATEGORY_SECTOR = {
    'Lithium Producers':        'industrial minerals',
    'Rare Earth Elements':      'industrial minerals',
    'Cobalt Producers':         'metals mining',
    'Copper Producers':         'metals mining',
    'Nickel Producers':         'metals mining',
    'Graphite Producers':       'industrial minerals',
    'Aluminum Producers':       'construction minerals',
    'Zinc Lead Tin Producers':  'metals mining',
    'Multi-Metal Refiners':     'metals mining',
    'Specialty Critical Materials': 'industrial minerals',
    'Other mines- recyclers':   'recycled aggregates',
}

DATA_COMPLETENESS_FLAGS = {
    "review_score": "manual_only",
    "defect_rate_ppm": "manual_only",
    "on_time_delivery_rate": "manual_only",
    "pricing": "api_only",
    "inventory_levels": "api_only",
}


def make_canonical(name: str) -> str:
    n = str(name).lower().strip()
    n = n.replace("&", "and")
    n = unicodedata.normalize("NFD", n)
    n = "".join(c for c in n if unicodedata.category(c) != "Mn")
    n = re.sub(r"[^\w\s-]", "", n)
    n = re.sub(r"\s+", "-", n.strip())
    return re.sub(r"-{2,}", "-", n).strip("-")


def infer_tier(type_str: str) -> int:
    if not type_str:
        return 1
    t = type_str.lower()
    if any(x in t for x in ["refiner", "smelter", "recycler", "processor"]):
        return 2
    if "trader" in t or "distributor" in t:
        return 3
    return 1


def infer_sector(products: str, sheet: str) -> str:
    base = CATEGORY_SECTOR.get(sheet, "metals mining")
    if not products:
        return base
    p = products.lower()
    KEYWORDS = [
        ("recycl", "recycled aggregates"),
        ("graphite", "industrial minerals"), ("silica", "industrial minerals"),
        ("potash", "industrial minerals"), ("fluorspar", "industrial minerals"),
        ("coal", "coal"), ("lithium", "metals mining"), ("cobalt", "metals mining"),
        ("nickel", "metals mining"), ("copper", "metals mining"),
        ("zinc", "metals mining"), ("rare earth", "metals mining"),
        ("aluminum", "metals mining"), ("aluminium", "metals mining"),
    ]
    for kw, sector in KEYWORDS:
        if kw in p:
            return sector
    return base


def clean_url(w) -> str | None:
    if not w:
        return None
    w = str(w).strip().split(",")[0].strip()
    if not w.startswith("http") and "." in w:
        w = "https://" + w
    return w if w.startswith("http") else None


def norm_transparency(t) -> str | None:
    if not t:
        return None
    t = str(t).strip()
    for level in ["High", "Medium-High", "Medium", "Low-Medium", "Low"]:
        if level.lower() in t.lower():
            return level
    return None


def parse_excel(xlsx_path: str) -> list[dict]:
    """Parse all companies from Excel — same logic as seed script."""
    wb = load_workbook(xlsx_path, read_only=True)
    companies = []

    for sheet_name in wb.sheetnames:
        if sheet_name == "Material web indexes":
            continue
        ws = wb[sheet_name]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]

        # Find header row
        header_idx = next(
            (i for i, r in enumerate(rows)
             if len(r) > 1 and r[1] and "Country" in str(r[1])),
            None
        )

        if sheet_name == "Other mines- recyclers":
            for row in rows:
                if not row or not row[0]:
                    continue
                name = str(row[0]).strip()
                if name.startswith("//") or name in ("Name", "Company Name"):
                    continue
                col1 = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                if col1.startswith("http") or (col1 and len(col1) > 20 and not col1.startswith("http")):
                    companies.append({
                        "company_name": name,
                        "canonical_name": make_canonical(name),
                        "website": col1 if col1.startswith("http") else None,
                        "headquarters_location": str(row[2]).strip() if not col1.startswith("http") and len(row) > 2 and row[2] else None,
                        "company_description": col1 if not col1.startswith("http") else None,
                        "industry_sector": "recycled aggregates",
                        "supply_chain_tier": 2,
                        "type_description": "Recycler",
                        "products_raw": None,
                        "data_transparency_level": None,
                        "sheet": sheet_name,
                    })

        elif header_idx is not None:
            for row in rows[header_idx + 1:]:
                if not row or not row[0] or str(row[0]).strip() in ("", "Company Name"):
                    continue
                name = str(row[0]).strip()
                type_str = str(row[2]).strip() if len(row) > 2 and row[2] else None
                products = str(row[3]).strip() if len(row) > 3 and row[3] else None
                website = clean_url(row[4] if len(row) > 4 else None)
                description = str(row[5]).strip() if len(row) > 5 and row[5] else None
                transparency = norm_transparency(row[6] if len(row) > 6 else None)
                hq = str(row[1]).strip() if len(row) > 1 and row[1] else None

                companies.append({
                    "company_name": name,
                    "canonical_name": make_canonical(name),
                    "headquarters_location": hq,
                    "website": website,
                    "company_description": description,
                    "type_description": type_str,
                    "products_raw": products,
                    "data_transparency_level": transparency,
                    "industry_sector": infer_sector(products or "", sheet_name),
                    "supply_chain_tier": infer_tier(type_str or ""),
                    "sheet": sheet_name,
                })

    print(f"  Parsed {len(companies)} companies from Excel")
    return companies


def build_patch_fields(c: dict) -> dict:
    """Build the complete set of fields to patch onto an existing record."""
    name = c["company_name"]
    website = c.get("website")
    products = c.get("products_raw") or ""
    canonical = make_canonical(name)

    return {
        # BGS base fields
        "supplier_id": None,
        "duns_number": None,
        "company_name": name,
        "canonical_name": canonical,
        "headquarters_location": c.get("headquarters_location"),
        "website": website,
        "company_description": c.get("company_description"),
        "industry_sector": c.get("industry_sector", "metals mining"),
        "supply_chain_tier": c.get("supply_chain_tier", 1),
        "typical_lead_time_days": None,
        "is_verified": False,
        "manufacturing_sites": [],
        "products_offered": [{
            "product_name": products,
            "grade": None,
            "product_id": f"{canonical[:15]}_01",
            "category": (c.get("sheet", "")).upper()[:20],
            "source_url": website,
            "datasheet_url": None,
            "cross_graph_material_id": None,
        }] if products else [],
        "certification_references": [],
        "certifications_raw": None,
        "regulation_references": [],
        "data_completeness_flags": DATA_COMPLETENESS_FLAGS.copy(),
        "sources": [{
            "source_name": " Critical Materials Intelligence — Intern Research 2025",
            "source_url": website or "https://.ai",
            "doi": None,
            "tier": "tier2",
        }],
        # Extras
        "extras": [{
            "type_description": c.get("type_description"),
            "primary_commodities_raw": products,
            "data_transparency_level": c.get("data_transparency_level"),
            "source_sheets": c.get("sheet", ""),
        }],
    }


class API:
    def __init__(self, base: str, email: str, password: str):
        self.base = base.rstrip("/")
        self.s = requests.Session()
        self.s.headers["Content-Type"] = "application/json"
        r = self.s.post(f"{self.base}/api/v1/auth/login",
                        json={"email": email, "password": password})
        r.raise_for_status()
        self.s.headers["Authorization"] = f"Bearer {r.json()['access_token']}"
        print(f"✓ Authenticated as {email}")

    def get(self, path, **kw):
        r = self.s.get(f"{self.base}{path}", **kw)
        r.raise_for_status()
        return r.json()

    def patch_record(self, source_id: str, record_id: str, fields: dict) -> dict:
        r = self.s.patch(
            f"{self.base}/api/v1/sources/{source_id}/records/{record_id}",
            json={"extracted_fields": fields}
        )
        r.raise_for_status()
        return r.json()


def run(args):
    print("\n[1/4] Parsing Excel…")
    xlsx = args.xlsx
    if not Path(xlsx).exists():
        xlsx = "Intern-_Critical_Materials_Suppliers.xlsx"
    if not Path(xlsx).exists():
        sys.exit(f"Excel not found: {args.xlsx}")
    companies = parse_excel(xlsx)

    # Build a lookup by canonical_name
    company_map = {c["canonical_name"]: c for c in companies}

    api = API(args.url, args.email, args.password)

    print("\n[2/4] Finding project…")
    projects = api.get("/api/v1/projects").get("items", [])
    project_name = args.project or "Critical Materials Intelligence"
    project = next((p for p in projects if project_name.lower() in p["name"].lower()), None)
    if not project:
        sys.exit(f'Project "{project_name}" not found in . Check --project name.')
    print(f"  Found: {project['name']} ({project['id']})")

    print("\n[3/4] Loading all sources in project…")
    sources_raw = api.get("/api/v1/sources", params={"project_id": project["id"]})
    sources = sources_raw if isinstance(sources_raw, list) else sources_raw.get("items", [])
    print(f"  Found {len(sources)} sources")

    print(f"\n[4/4] Patching records{'  [DRY RUN]' if args.dry_run else ''}…")
    patched = 0
    skipped = 0
    not_found = 0

    for source in sources:
        source_id = source["id"]
        source_name = source["name"]
        cn = make_canonical(source_name)

        # Find matching company in Excel
        company = company_map.get(cn)
        if not company:
            # Try fuzzy: source name might differ slightly
            for key, c in company_map.items():
                if key[:10] == cn[:10]:
                    company = c
                    break

        if not company:
            not_found += 1
            continue

        # Get existing records for this source
        try:
            records_data = api.get(f"/api/v1/sources/{source_id}/records",
                                   params={"page_size": 5})
            records = records_data.get("items", records_data) if isinstance(records_data, dict) else records_data
        except Exception:
            records = []

        if not records:
            skipped += 1
            continue

        # Patch the first record (each source has exactly 1 record)
        record = records[0]
        record_id = record["id"]
        patch_fields = build_patch_fields(company)

        if args.dry_run:
            print(f"  [DRY] {source_name[:50]:50s} → would patch {len(patch_fields)} fields")
            patched += 1
        else:
            try:
                api.patch_record(source_id, record_id, patch_fields)
                patched += 1
                if patched % 20 == 0:
                    print(f"  Patched {patched}/{len(sources)}…")
                time.sleep(0.1)
            except Exception as e:
                print(f"  ✗ {source_name}: {e}")

    print(f"\n✓ Done!")
    print(f"  Patched:   {patched}")
    print(f"  Skipped:   {skipped} (no existing records)")
    print(f"  Not found: {not_found} (company not in Excel)")
    if args.dry_run:
        print("\n  [DRY RUN] No changes were saved. Remove --dry-run to apply.")
    else:
        print("\n  Open  → refresh to see updated records.")


def main():
    p = argparse.ArgumentParser(
        description="Patch existing  records with correct Excel data — no duplicates created"
    )
    p.add_argument("--url", default="https://-platform-production.up.railway.app")
    p.add_argument("--email", required=True)
    p.add_argument("--password", required=True)
    p.add_argument("--xlsx", default="Intern- Critical_Materials_Suppliers.xlsx",
                   help="Path to the Excel file")
    p.add_argument("--project", default="Critical Materials Intelligence",
                   help=" project name to patch")
    p.add_argument("--dry-run", action="store_true",
                   help="Show what would change without saving anything")
    args = p.parse_args()
    run(args)


if __name__ == "__main__":
    main()
