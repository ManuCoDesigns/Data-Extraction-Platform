"""
Sources API — the Kanban-tracked dataset workflow.

A Source is one tracked dataset within a project: it has a schema, an
optional source website, an assigned extractor, an assigned reviewer, and
moves through a status pipeline as work happens (see SourceStatus enum).

Upload flow (synchronous, no Celery dependency):
  1. Extractor uploads a CSV/Excel/JSON file of already-extracted rows
  2. Each row is mapped onto the schema's fields and validated structurally
  3. Records are created with is_schema_valid + validation_errors set
  4. Source status moves to NEEDS_FIXES (if any invalid) or READY_FOR_REVIEW

Re-uploading replaces the source's current record set (simple, predictable —
no row-level merge/dedup in this version). Individual records can also be
fixed inline via PATCH without a full re-upload.
"""
import io, json, math, re
from datetime import datetime, timezone
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
import pandas as pd

from app.db.session import get_db
from app.core.security import get_current_user
from app.models.all_models import (
    Source, SourceStatus, Project, ProjectMember, User, Schema, SchemaVersion,
    ExtractionJob, ExtractedRecord, JobStatus, SourceType as FileSourceType,
    ExtractionConfidence, ReviewStatus, AuditLog, AuditAction, Notification,
    LLMCallLog, SubmissionBatch, UploadedFileEntry,
)
from app.schemas.api_schemas import (
    SourceCreate, SourceUpdate, SourceOut, SourceUploadSummary,
    SourceRecordFix, SourceRecordReview, RecordOut, PaginatedResponse,
)
from app.services.schema_validator import validate_record, map_row_to_fields

router = APIRouter(prefix="/sources", tags=["sources"])

ALLOWED_UPLOAD_EXTENSIONS = {".csv", ".xlsx", ".xls", ".json", ".pdf", ".txt", ".zip"}

# The backend currently reads the whole upload into memory before parsing
# it (pandas/json need the full content anyway). A 126MB file was enough to
# get the Railway container OOM-killed mid-request (502, "connection closed
# unexpectedly") since parsing overhead runs several times the raw file
# size. Reject oversized uploads immediately, before any memory is spent,
# with a clear message — instead of a silent 2-minute crash.
MAX_UPLOAD_SIZE_BYTES = 60 * 1024 * 1024  # 60MB per file
# Extensions that go through AI extraction (not mechanical row mapping)
AI_EXTRACTION_EXTENSIONS = {".pdf", ".txt"}

# ─── Folder-upload file classification ─────────────────────────────────────
# Real SOP batch deliveries have used at least three different folder/naming
# conventions in practice (the full "materials/suppliers/.../qa_ready_batch"
# structure, an abbreviated "compliance/qa/sources" structure, and a
# "suppliers/evidence" structure) — so classification can't rely on a fixed
# folder-name list alone. This uses three signals, most reliable first:
#
#   1. Unambiguous primary-data folder names — never used for anything else.
#   2. Filename keywords (qa_checklist, review_log, source list/index/manifest)
#      — checked BEFORE generic folder names, since a filename like
#      "row161_review_log.txt" is reliable even inside a shared/ambiguous
#      folder such as "qa/" that holds more than one file role.
#   3. Unambiguous full folder names, as a fallback if the filename alone
#      didn't indicate a role.
#
# Anything that still doesn't match is only treated as primary data if it's
# unambiguous: either it sits at the flat root of the upload (a single-file
# delivery with no folder at all), or its name matches the batch's own slug
# (e.g. "45-8-energy.json" inside "batch_20260819_45-8-energy/..."). A file
# that matches NONE of this — a stray note, an old backup, anything not
# actually part of the delivery — is skipped entirely rather than guessed
# into becoming fake data. This is the fix for folder uploads silently
# ingesting unrelated files as bogus records.
PRIMARY_DATA_FOLDERS = {"materials", "suppliers", "compliance_records", "research"}
MANIFEST_FOLDER_NAMES = {"source_lists"}
QA_CHECKLIST_FOLDER_NAMES = {"qa_checklists"}
REVIEW_LOG_FOLDER_NAMES = {"escalation_notes"}


def _derive_batch_slugs(paths: list[str]) -> set[str]:
    """
    Extracts likely 'source slug' candidates from the batch's root folder
    name(s) present across the uploaded paths — e.g. both the full root
    "batch_20260819_45-8-energy" and, with the date prefix stripped,
    "45-8-energy". Used only as the last-resort signal for classifying an
    otherwise-unrecognised file.
    """
    roots = set()
    for p in paths:
        parts = [x for x in p.replace("\\", "/").split("/") if x]
        if len(parts) > 1:
            roots.add(parts[0].lower())
    slugs = set()
    for root in roots:
        slugs.add(root)
        m = re.match(r"^batch_\d{8}_(.+)$", root)
        if m:
            slugs.add(m.group(1))
    return slugs


def _normalize_slug(s: str) -> str:
    return re.sub(r"[-_]+", "-", s.strip().lower()).strip("-")


def _classify_batch_file(full_path: str, batch_slugs: set[str] | None = None) -> str:
    """
    Classifies one uploaded file by its ROLE in a folder/ZIP upload.
    Returns one of: "primary_data" | "manifest" | "qa_checklist" | "review_log" | "unknown"

    Only "primary_data" files are parsed into ExtractedRecords. Everything
    else is either recognised-but-not-data (manifest/checklist/review log)
    or "unknown" — and "unknown" files are skipped entirely, never ingested.
    """
    parts = [p for p in full_path.replace("\\", "/").split("/") if p]
    parts_lower = [p.lower() for p in parts]
    fname_lower = parts[-1].lower() if parts else full_path.lower()
    stem = fname_lower.rsplit(".", 1)[0] if "." in fname_lower else fname_lower
    ext = ("." + fname_lower.rsplit(".", 1)[1]) if "." in fname_lower else ""

    # 1. Unambiguous primary-data folders
    if any(p in PRIMARY_DATA_FOLDERS for p in parts_lower):
        return "primary_data"

    # 2. Filename keywords — checked before generic folder names
    if "qa_checklist" in stem or "qa-checklist" in stem or "qachecklist" in stem:
        return "qa_checklist"
    if "review_log" in stem or "review-log" in stem or "escalation" in stem:
        return "review_log"
    if (stem.endswith("_sources") or stem.endswith("-sources")
            or "source_list" in stem or "sourcelist" in stem or "source-list" in stem
            or "source_index" in stem or "sourceindex" in stem or "manifest" in stem):
        return "manifest"

    # 3. Unambiguous full folder names, as a fallback
    if any(p in MANIFEST_FOLDER_NAMES for p in parts_lower):
        return "manifest"
    if any(p in QA_CHECKLIST_FOLDER_NAMES for p in parts_lower):
        return "qa_checklist"
    if any(p in REVIEW_LOG_FOLDER_NAMES for p in parts_lower):
        return "review_log"

    # 4. Default to primary_data ONLY when unambiguous
    if ext in {".json", ".csv", ".xlsx", ".xls"}:
        is_at_root = len(parts) == 1
        if is_at_root:
            return "primary_data"
        if batch_slugs:
            stem_norm = _normalize_slug(stem)
            for slug in batch_slugs:
                if stem_norm == _normalize_slug(slug):
                    return "primary_data"

    return "unknown"


# ─── Permission helpers ──────────────────────────────────────────────────────

def _project_role(user: User, project: Project) -> str | None:
    user_roles = {r.role.value for r in user.roles}
    if "org_admin" in user_roles:
        return "org_admin"
    m = next((m for m in project.members if m.user_id == user.id), None)
    return m.role.value if m else None


def _is_org_admin(user: User) -> bool:
    return "org_admin" in {r.role.value for r in user.roles}


def _user_roles(user: User) -> set:
    return {r.role.value for r in user.roles}


def _can_access(user: User, project: Project) -> bool:
    if _is_org_admin(user):
        return True
    return _project_role(user, project) is not None


def _is_project_admin(user: User, project: Project) -> bool:
    if _is_org_admin(user):
        return True
    return _project_role(user, project) in ("org_admin", "project_admin")


def _can_manage_source(user: User, source: Source) -> bool:
    """Admin always wins. For non-admins, check project membership via project_id (no lazy-load)."""
    if _is_org_admin(user):
        return True
    roles = _user_roles(user)
    if "project_admin" in roles:
        return True
    return False


def _is_assigned_extractor(user: User, source: Source) -> bool:
    """Admin and project_admin can act as extractor on any source."""
    if _is_org_admin(user):
        return True
    roles = _user_roles(user)
    if "project_admin" in roles:
        return True
    return source.assigned_extractor_id == user.id


def _is_assigned_reviewer(user: User, source: Source) -> bool:
    """Admin, project_admin, and qa_lead can review any source."""
    if _is_org_admin(user):
        return True
    roles = _user_roles(user)
    if "project_admin" in roles or "qa_lead" in roles:
        return True
    return source.assigned_reviewer_id == user.id


# ─── Serialization ───────────────────────────────────────────────────────────

def _serialize_source(s: Source) -> SourceOut:
    return SourceOut(
        id=s.id, project_id=s.project_id, schema_id=s.schema_id,
        schema_name=s.schema.name if s.schema else None,
        name=s.name, description=s.description, website_url=s.website_url,
        category=getattr(s, "category", None),
        country=getattr(s, "country", None),
        type=getattr(s, "type", None),
        external_ref_id=getattr(s, "external_ref_id", None),
        external_system=getattr(s, "external_system", None),
        status=s.status.value,
        assigned_extractor_id=s.assigned_extractor_id,
        assigned_extractor_name=s.extractor.full_name if s.extractor else None,
        assigned_reviewer_id=s.assigned_reviewer_id,
        assigned_reviewer_name=s.reviewer.full_name if s.reviewer else None,
        total_records=s.total_records or 0, valid_records=s.valid_records or 0,
        invalid_records=s.invalid_records or 0, approved_records=s.approved_records or 0,
        notes=s.notes, created_at=s.created_at, updated_at=s.updated_at,
        extraction_started_at=s.extraction_started_at,
        extraction_completed_at=s.extraction_completed_at,
        llm_verification_started_at=getattr(s, "llm_verification_started_at", None),
        llm_verification_completed_at=getattr(s, "llm_verification_completed_at", None),
        review_started_at=s.review_started_at, review_completed_at=s.review_completed_at,
        approved_at=s.approved_at, created_by=s.created_by,
        reset_count=getattr(s, "reset_count", 0) or 0,
    )


def _serialize_record(r: ExtractedRecord) -> RecordOut:
    return RecordOut(
        is_escalation_only=getattr(r, "is_escalation_only", False),
        escalation_reason=getattr(r, "escalation_reason", None),
        id=r.id, job_id=r.job_id, schema_version=r.schema_version,
        extraction_confidence=r.extraction_confidence.value,
        pipeline_warnings=r.pipeline_warnings or [],
        is_schema_valid=r.is_schema_valid, validation_errors=r.validation_errors or [],
        review_status=r.review_status.value, review_note=r.review_note,
        reviewed_by=r.reviewed_by, reviewed_at=r.reviewed_at,
        llm_verdict=r.llm_verdict.value if r.llm_verdict else None,
        llm_confidence=r.llm_confidence, llm_field_flags=r.llm_field_flags or [],
        llm_reason=r.llm_reason, llm_skipped=r.llm_skipped,
        web_verified=r.web_verified,
        web_check_flags=r.web_check_flags or [],
        web_check_summary=r.web_check_summary,
        extracted_fields=r.extracted_fields or {}, raw_text=r.raw_text or "",
        is_submitted=r.is_submitted, canonical_name=r.canonical_name,
        created_at=r.created_at,
        correction_count=getattr(r, "correction_count", 0) or 0,
        revision_count=getattr(r, "revision_count", 0) or 0,
        reviewer_field_comments=getattr(r, "reviewer_field_comments", None) or {},
        admin_review_note=getattr(r, "admin_review_note", None),
        admin_reviewed_at=getattr(r, "admin_reviewed_at", None),
    )


def _get_source_or_404(source_id: str, db: Session) -> Source:
    source = db.query(Source).filter(Source.id == source_id).first()
    if not source:
        raise HTTPException(status_code=404, detail="Source not found")
    return source


def _safe_clear_old_jobs(source_id: str, db: Session) -> int:
    """
    Deletes every ExtractionJob (and its ExtractedRecords) for a source.

    audit_log.record_id / job_id / source_id are no longer enforced as
    database-level foreign keys (migration 015) — they were dropped after
    an in-place ALTER (migration 014) left Postgres's internal RI trigger
    for this constraint permanently corrupted, breaking every delete that
    touched extracted_records/extraction_jobs regardless of how the delete
    was written. With the constraint gone, plain bulk deletes work again.
    We still null out audit_log references first, purely for data hygiene
    (so old audit rows don't point at IDs that no longer exist).
    """
    job_ids = [j.id for j in db.query(ExtractionJob).filter(ExtractionJob.source_id == source_id).all()]
    if not job_ids:
        return 0

    record_ids = [r.id for r in db.query(ExtractedRecord.id).filter(ExtractedRecord.job_id.in_(job_ids)).all()]
    if record_ids:
        db.query(AuditLog).filter(AuditLog.record_id.in_(record_ids)).update(
            {"record_id": None}, synchronize_session=False
        )
    db.query(AuditLog).filter(AuditLog.job_id.in_(job_ids)).update(
        {"job_id": None}, synchronize_session=False
    )
    db.query(LLMCallLog).filter(LLMCallLog.job_id.in_(job_ids)).delete(synchronize_session=False)
    db.query(SubmissionBatch).filter(SubmissionBatch.job_id.in_(job_ids)).update(
        {"job_id": None}, synchronize_session=False
    )
    deleted = db.query(ExtractedRecord).filter(ExtractedRecord.job_id.in_(job_ids)).delete(synchronize_session=False)
    db.query(ExtractionJob).filter(ExtractionJob.id.in_(job_ids)).delete(synchronize_session=False)
    return deleted



def _recompute_counts(source: Source, db: Session):
    """
    Only counts records from the MOST RECENT extraction job for this source.
    Defensive by design: if an old job ever fails to get cleaned up on
    re-upload (e.g. a stale FK constraint blocking deletion), stale records
    from it must never silently inflate the visible count again.
    """
    latest_job = (
        db.query(ExtractionJob)
        .filter(ExtractionJob.source_id == source.id)
        .order_by(ExtractionJob.created_at.desc())
        .first()
    )
    if not latest_job:
        source.total_records = 0
        source.valid_records = 0
        source.invalid_records = 0
        source.approved_records = 0
        return

    records = db.query(ExtractedRecord).filter(ExtractedRecord.job_id == latest_job.id).all()
    source.total_records = len(records)
    source.valid_records = sum(1 for r in records if r.is_schema_valid)
    source.invalid_records = source.total_records - source.valid_records
    source.approved_records = sum(1 for r in records if r.review_status == ReviewStatus.APPROVED)


# ─── CRUD ────────────────────────────────────────────────────────────────────

@router.get("", response_model=list[SourceOut])
def list_sources(
    project_id: str = Query(None),
    status: str = Query(None),
    assigned_to_me: bool = Query(False, description="Only sources where I'm the extractor or reviewer"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    user_roles = {r.role.value for r in current_user.roles}
    is_admin = "org_admin" in user_roles

    if project_id:
        project = db.query(Project).filter(Project.id == project_id, Project.deleted_at == None).first()
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")
        if not _can_access(current_user, project):
            raise HTTPException(status_code=403, detail="Access denied")
        q = db.query(Source).filter(Source.project_id == project_id)
    else:
        if is_admin:
            q = db.query(Source)
        else:
            accessible_project_ids = [
                m.project_id for m in db.query(ProjectMember).filter(ProjectMember.user_id == current_user.id).all()
            ]
            if not accessible_project_ids:
                return []
            q = db.query(Source).filter(Source.project_id.in_(accessible_project_ids))

    if status:
        try:
            q = q.filter(Source.status == SourceStatus(status))
        except ValueError:
            pass

    if assigned_to_me:
        q = q.filter(
            (Source.assigned_extractor_id == current_user.id) | (Source.assigned_reviewer_id == current_user.id)
        )

    sources = q.order_by(Source.updated_at.desc()).all()
    return [_serialize_source(s) for s in sources]


@router.post("", response_model=SourceOut, status_code=201)
def create_source(
    payload: SourceCreate,
    project_id: str = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    project = db.query(Project).filter(Project.id == project_id, Project.deleted_at == None).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if not _is_project_admin(current_user, project):
        raise HTTPException(status_code=403, detail="Only project admins can create sources")

    if payload.schema_id:
        schema = db.query(Schema).filter(Schema.id == payload.schema_id, Schema.project_id == project_id).first()
        if not schema:
            raise HTTPException(status_code=404, detail="Schema not found in this project")
        if not db.query(SchemaVersion).filter(SchemaVersion.schema_id == schema.id).first():
            raise HTTPException(status_code=422, detail="Schema has no versions — add fields first")

    source = Source(
        project_id=project_id, schema_id=payload.schema_id, name=payload.name,
        description=payload.description, website_url=payload.website_url,
        category=payload.category,
        country=payload.country,
        type=payload.type,
        assigned_extractor_id=payload.assigned_extractor_id,
        assigned_reviewer_id=payload.assigned_reviewer_id,
        status=SourceStatus.EXTRACTING if payload.assigned_extractor_id else SourceStatus.NOT_STARTED,
        created_by=current_user.id,
    )
    db.add(source)
    db.flush()
    db.add(AuditLog(
        user_id=current_user.id, project_id=project_id,
        action=AuditAction.SOURCE_CREATED, after_value={"name": payload.name},
    ))
    db.commit()
    db.refresh(source)
    return _serialize_source(source)


# ─── Live Team Workload ──────────────────────────────────────────────────────

@router.get("/workload")
def team_workload(
    project_id: str | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Live "who is handling what" board — every source that isn't fully
    delivered yet, with its current extractor/reviewer, how long each has
    had it, and where it's stuck. Powers the Team Workload page.
    """
    q = db.query(Source).filter(Source.status != SourceStatus.APPROVED)
    if project_id:
        q = q.filter(Source.project_id == project_id)
    sources = q.order_by(Source.updated_at.desc()).all()

    now = datetime.now(timezone.utc)
    user_cache: dict[str, str] = {}

    def user_name(uid: str | None) -> str | None:
        if not uid:
            return None
        if uid not in user_cache:
            u = db.query(User).filter(User.id == uid).first()
            user_cache[uid] = u.full_name if u else "Unknown"
        return user_cache[uid]

    def elapsed_label(started: datetime | None) -> str | None:
        if not started:
            return None
        delta = now - started
        hrs = delta.total_seconds() / 3600
        if hrs < 1:
            return f"{int(delta.total_seconds() / 60)}m"
        if hrs < 24:
            return f"{hrs:.1f}h"
        return f"{delta.days}d {int(hrs % 24)}h"

    projects_cache: dict[str, str] = {}
    def project_name(pid: str) -> str:
        if pid not in projects_cache:
            p = db.query(Project).filter(Project.id == pid).first()
            projects_cache[pid] = p.name if p else "Unknown Project"
        return projects_cache[pid]

    STAGE_LABELS = {
        "not_started":       "Not Started",
        "extracting":        "Extraction",
        "needs_fixes":       "Extraction (Fixing Errors)",
        "llm_verification":  "LLM Verification",
        "ready_for_review":  "Ready for Human Review",
        "in_review":         "Human Review",
        "changes_requested": "Extraction (Corrections)",
        "approved":          "Delivered",
    }

    rows = []
    for s in sources:
        extractor = user_name(s.assigned_extractor_id)
        reviewer = user_name(s.assigned_reviewer_id)

        is_extracting = s.status in (SourceStatus.EXTRACTING, SourceStatus.NEEDS_FIXES, SourceStatus.CHANGES_REQUESTED)
        is_llm_verifying = s.status == SourceStatus.LLM_VERIFICATION
        is_reviewing = s.status in (SourceStatus.IN_REVIEW, SourceStatus.READY_FOR_REVIEW)

        llm_started = getattr(s, "llm_verification_started_at", None)

        is_waiting_for_review = s.status == SourceStatus.READY_FOR_REVIEW and not s.review_started_at
        waiting_since = getattr(s, "llm_verification_completed_at", None) or s.extraction_completed_at

        rows.append({
            "source_id": s.id,
            "source_name": s.name,
            "project_id": s.project_id,
            "project_name": project_name(s.project_id),
            "status": s.status.value,
            "pipeline_stage": STAGE_LABELS.get(s.status.value, s.status.value),
            "extractor": extractor,
            "extractor_elapsed": elapsed_label(s.extraction_started_at) if is_extracting else None,
            "llm_verifying": is_llm_verifying,
            "llm_elapsed": elapsed_label(llm_started) if is_llm_verifying else None,
            "reviewer": reviewer,
            "reviewer_elapsed": elapsed_label(s.review_started_at) if is_reviewing and s.review_started_at else None,
            "waiting_for_review": is_waiting_for_review,
            "waiting_elapsed": elapsed_label(waiting_since) if is_waiting_for_review else None,
            "total_records": s.total_records or 0,
            "valid_records": s.valid_records or 0,
            "invalid_records": s.invalid_records or 0,
            "approved_records": s.approved_records or 0,
            "updated_at": s.updated_at.isoformat() if s.updated_at else None,
            "extraction_started_at": s.extraction_started_at.isoformat() if s.extraction_started_at else None,
            "unclaimed": not s.assigned_extractor_id,
        })

    by_person: dict[str, dict] = {}
    for r in rows:
        for role_key, name_key, elapsed_key in [
            ("extractor", "extractor", "extractor_elapsed"),
            ("reviewer", "reviewer", "reviewer_elapsed"),
        ]:
            name = r[name_key]
            if not name:
                continue
            if name not in by_person:
                by_person[name] = {"name": name, "extracting": 0, "reviewing": 0}
            if role_key == "extractor" and r[elapsed_key]:
                by_person[name]["extracting"] += 1
            if role_key == "reviewer" and r[elapsed_key]:
                by_person[name]["reviewing"] += 1

    return {
        "sources": rows,
        "by_person": list(by_person.values()),
        "unclaimed_count": len([r for r in rows if r["unclaimed"]]),
        "llm_verifying_count": len([r for r in rows if r["llm_verifying"]]),
        "generated_at": now.isoformat(),
    }


# ─── Escalations — records sent back with feedback ────────────────────────────

@router.get("/escalations")
def list_escalations(
    project_id: str | None = None,
    mine_only: bool = True,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Every record currently sitting with the extractor because it was sent
    back — either by a reviewer (reject) or by an admin (return for
    correction) — along with the actual feedback message and who wrote it.

    mine_only=True (default) scopes to sources where the current user is
    the assigned extractor. Admins/reviewers can pass mine_only=false to
    see everything across a project.
    """
    user_roles = {r.role.value for r in current_user.roles}
    is_admin = "org_admin" in user_roles or "project_admin" in user_roles

    q = db.query(ExtractedRecord).join(
        ExtractionJob, ExtractedRecord.job_id == ExtractionJob.id
    ).join(
        Source, ExtractionJob.source_id == Source.id
    ).filter(
        ExtractedRecord.review_status.in_([ReviewStatus.REJECTED, ReviewStatus.PENDING]),
        ExtractedRecord.correction_count > 0,
    )

    if project_id:
        q = q.filter(Source.project_id == project_id)
    if mine_only and not is_admin:
        q = q.filter(Source.assigned_extractor_id == current_user.id)
    elif mine_only:
        q = q.filter(Source.assigned_extractor_id == current_user.id)

    records = q.order_by(ExtractedRecord.updated_at.desc()).all()

    projects_cache: dict[str, str] = {}
    def project_name(pid: str) -> str:
        if pid not in projects_cache:
            p = db.query(Project).filter(Project.id == pid).first()
            projects_cache[pid] = p.name if p else "Unknown Project"
        return projects_cache[pid]

    escalations = []
    for r in records:
        job = db.query(ExtractionJob).filter(ExtractionJob.id == r.job_id).first()
        source = db.query(Source).filter(Source.id == job.source_id).first() if job else None
        if not source:
            continue

        latest_message = ""
        latest_by = ""
        latest_role = ""
        latest_ts = None
        comments = r.reviewer_field_comments or {}
        all_entries = []
        for field, entries in comments.items():
            for e in (entries or []):
                if e.get("type") in ("correction", "rejection"):
                    all_entries.append({**e, "field": field})
        if all_entries:
            all_entries.sort(key=lambda e: e.get("ts", ""))
            last = all_entries[-1]
            latest_message = last.get("comment", "")
            latest_by = last.get("user", "")
            latest_role = last.get("role", "")
            latest_ts = last.get("ts")
        elif r.review_note:
            latest_message = r.review_note

        record_label = None
        if isinstance(r.extracted_fields, dict):
            record_label = r.extracted_fields.get("_source_file")
            if record_label:
                record_label = record_label.split("/")[-1]
        record_label = record_label or (r.canonical_name or r.id[:8])

        escalations.append({
            "record_id": r.id,
            "source_id": source.id,
            "source_name": source.name,
            "project_id": source.project_id,
            "project_name": project_name(source.project_id),
            "record_label": record_label,
            "correction_count": r.correction_count or 0,
            "message": latest_message,
            "by": latest_by,
            "role": latest_role,
            "when": latest_ts or (r.updated_at.isoformat() if r.updated_at else None),
            "all_messages": all_entries,
        })

    return {
        "escalations": escalations,
        "count": len(escalations),
    }


@router.get("/{source_id}", response_model=SourceOut)
def get_source(source_id: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    source = _get_source_or_404(source_id, db)
    if not _can_access(current_user, source.project):
        raise HTTPException(status_code=403, detail="Access denied")
    _recompute_counts(source, db)
    db.commit()
    return _serialize_source(source)


@router.patch("/{source_id}", response_model=SourceOut)
def update_source(
    source_id: str, payload: SourceUpdate,
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user),
):
    source = _get_source_or_404(source_id, db)
    if not _can_manage_source(current_user, source):
        raise HTTPException(status_code=403, detail="Only project admins can edit sources")

    before_status = source.status.value
    if payload.name is not None:
        source.name = payload.name
    if payload.description is not None:
        source.description = payload.description
    if payload.website_url is not None:
        source.website_url = payload.website_url
    if payload.category is not None:
        source.category = payload.category
    if payload.country is not None:
        source.country = payload.country
    if payload.type is not None:
        source.type = payload.type
    if payload.notes is not None:
        source.notes = payload.notes

    if payload.assigned_extractor_id is not None:
        source.assigned_extractor_id = payload.assigned_extractor_id or None
        if source.status == SourceStatus.NOT_STARTED and source.assigned_extractor_id:
            source.status = SourceStatus.EXTRACTING
            source.extraction_started_at = datetime.now(timezone.utc)
    if payload.assigned_reviewer_id is not None:
        source.assigned_reviewer_id = payload.assigned_reviewer_id or None

    if payload.status is not None:
        try:
            new_status = SourceStatus(payload.status)
        except ValueError:
            raise HTTPException(status_code=422, detail=f"Invalid status: {payload.status}")
        source.status = new_status
        if new_status == SourceStatus.IN_REVIEW and not source.review_started_at:
            source.review_started_at = datetime.now(timezone.utc)
        if new_status == SourceStatus.APPROVED:
            source.approved_at = datetime.now(timezone.utc)
            source.review_completed_at = source.review_completed_at or datetime.now(timezone.utc)

    db.add(AuditLog(
        user_id=current_user.id, project_id=source.project_id,
        action=AuditAction.SOURCE_STATUS_CHANGED,
        before_value={"status": before_status}, after_value={"status": source.status.value},
    ))
    db.commit()
    db.refresh(source)
    return _serialize_source(source)


# ─── Upload + validate ───────────────────────────────────────────────────────

@router.post("/{source_id}/upload", response_model=SourceUploadSummary)
async def upload_to_source(
    source_id: str,
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Two paths depending on file type:

    Structured (CSV / Excel / JSON):
      Rows are mapped mechanically onto schema fields and validated.
      Fast — no LLM call, results in seconds.

    Unstructured (PDF / TXT):
      Claude reads the document, understands the schema, and extracts
      every record it can find. Runs synchronously — no Celery needed.
      Takes 10–30 seconds depending on document size.
    """
    source = _get_source_or_404(source_id, db)
    if not _is_assigned_extractor(current_user, source):
        raise HTTPException(status_code=403, detail="Only the assigned extractor, project admin, or org admin can upload to this source")

    content_length = request.headers.get("content-length")
    if content_length and int(content_length) > MAX_UPLOAD_SIZE_BYTES:
        size_mb = int(content_length) / (1024 * 1024)
        limit_mb = MAX_UPLOAD_SIZE_BYTES / (1024 * 1024)
        raise HTTPException(
            status_code=413,
            detail=f"File is {size_mb:.1f}MB — the limit is {limit_mb:.0f}MB. "
                    f"Split it into smaller files, or use folder upload to send several at once.",
        )

    import os as _os
    filename = file.filename or ""
    ext = _os.path.splitext(filename)[1].lower()
    if not ext:
        ct = (file.content_type or "").lower()
        if "pdf" in ct:
            ext = ".pdf"
        elif "text" in ct or "plain" in ct:
            ext = ".txt"
        elif "csv" in ct:
            ext = ".csv"
        elif "json" in ct:
            ext = ".json"
        elif "sheet" in ct or "excel" in ct or "spreadsheet" in ct:
            ext = ".xlsx"

    if ext not in ALLOWED_UPLOAD_EXTENSIONS:
        raise HTTPException(status_code=422, detail=f"Unsupported file type '{ext or file.content_type}'. Accepted: PDF, TXT, CSV, XLSX, JSON.")

    schema_ver = db.query(SchemaVersion).filter(
        SchemaVersion.schema_id == source.schema_id
    ).order_by(SchemaVersion.version.desc()).first()

    schema_fields = []
    if schema_ver and schema_ver.definition:
        schema_fields = schema_ver.definition.get("fields", [])
    content = await file.read()

    if ext in AI_EXTRACTION_EXTENSIONS:
        try:
            _schema_def = schema_ver.definition if schema_ver else {"flexible_validation": True, "fields": []}
            rows = await _extract_with_llm(content, ext, _schema_def, source)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"AI extraction failed: {str(e)}")
        if not rows:
            raise HTTPException(status_code=422, detail="AI extraction found no records in the document. Check that the document matches the schema and try again.")
        extraction_method = "llm"
        file_breakdown: list[dict] = []
        files_processed = 1
        raw_entries = [{"relative_path": file.filename or "upload", "is_directory": False, "content": content}]

    elif ext == ".zip":
        rows, file_breakdown = _parse_zip(content)
        if not rows:
            skipped = [f for f in file_breakdown if f.get("error")]
            detail = f"No records found in ZIP. {len(file_breakdown)} file(s) checked."
            if skipped:
                detail += f" Errors: {'; '.join(f['filename'] + ': ' + f['error'] for f in skipped[:3])}"
            raise HTTPException(status_code=422, detail=detail)
        extraction_method = "structured"
        files_processed = len([f for f in file_breakdown if not f.get("error")])
        raw_entries = _read_raw_zip_entries(content)

    else:
        rows = _parse_rows(content, ext, file.filename or "")
        if not rows:
            raise HTTPException(status_code=422, detail="No rows found in the uploaded file.")
        extraction_method = "structured"
        file_breakdown = []
        files_processed = 1
        raw_entries = [{"relative_path": file.filename or "upload", "is_directory": False, "content": content}]

    return _finalize_upload(
        source=source, source_id=source_id, rows=rows, ext=ext,
        extraction_method=extraction_method, file_breakdown=file_breakdown,
        files_processed=files_processed, display_filename=file.filename or "upload",
        content_len=len(content), schema_ver=schema_ver, schema_fields=schema_fields,
        current_user=current_user, db=db, raw_entries=raw_entries,
    )


def _finalize_upload(
    source: "Source", source_id: str, rows: list[dict], ext: str,
    extraction_method: str, file_breakdown: list[dict], files_processed: int,
    display_filename: str, content_len: int,
    schema_ver, schema_fields: list[dict],
    current_user: "User", db: Session,
    raw_entries: list[dict] | None = None,
) -> "SourceUploadSummary":
    """
    Shared logic for turning parsed rows into an ExtractionJob + ExtractedRecords.
    Used by both single-file upload and multi-file/folder upload — keeps the
    schema-mapping, field-inference, and validation logic in exactly one place.

    raw_entries, if given, is a flat list of {relative_path, is_directory,
    content} dicts capturing EVERY file (and, for ZIP uploads, every empty
    directory) from the original upload verbatim — independent of whatever
    got parsed into rows. These are persisted once the job exists, so the
    exact original structure can always be browsed/re-downloaded later via
    GET /sources/{id}/files, regardless of what was classified as data.
    """
    _safe_clear_old_jobs(source_id, db)

    file_ext_type = {
        ".csv": FileSourceType.CSV, ".xlsx": FileSourceType.EXCEL,
        ".xls": FileSourceType.EXCEL, ".json": FileSourceType.CSV,
        ".pdf": FileSourceType.PDF, ".txt": FileSourceType.CSV,
    }.get(ext, FileSourceType.CSV)

    job = ExtractionJob(
        project_id=source.project_id, source_id=source_id,
        schema_id=source.schema_id, schema_version=(schema_ver.version if schema_ver else None),
        name=f"{source.name} — {extraction_method} upload {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')}",
        source_file_name=display_filename, source_file_size_bytes=content_len,
        source_type=file_ext_type, status=JobStatus.READY_FOR_REVIEW,
        total_raw_records=len(rows), total_extracted=len(rows),
        created_by=current_user.id,
    )
    db.add(job)
    db.flush()

    if raw_entries:
        for entry in raw_entries:
            db.add(UploadedFileEntry(
                job_id=job.id,
                relative_path=entry["relative_path"],
                is_directory=entry["is_directory"],
                content=entry["content"],
                size_bytes=len(entry["content"]) if entry.get("content") else 0,
            ))

    valid_count = 0
    for row in rows:
        if extraction_method == "llm":
            mapped = {k: v for k, v in row.items() if k != "_raw_text"}
            raw_text = row.get("_raw_text", json.dumps(row, ensure_ascii=False, default=str))
        elif ext in (".json", ".zip"):
            mapped = {k: v for k, v in row.items()}
            for field_def in schema_fields:
                if "fixed_value" in field_def:
                    mapped[field_def["name"]] = field_def["fixed_value"]
            raw_text = json.dumps(row, ensure_ascii=False, default=str)
        else:
            mapped = map_row_to_fields(row, schema_fields)
            raw_text = json.dumps(row, ensure_ascii=False, default=str)

            if not mapped.get("canonical_name") and mapped.get("company_name"):
                import unicodedata as _ud
                n = str(mapped["company_name"]).lower().strip()
                n = n.replace("&", "and")
                n = _ud.normalize("NFD", n)
                n = "".join(c for c in n if _ud.category(c) != "Mn")
                n = __import__("re").sub(r"[^\w\s-]", "", n)
                n = __import__("re").sub(r"\s+", "-", n.strip())
                n = __import__("re").sub(r"-{2,}", "-", n)
                mapped["canonical_name"] = n

            if not mapped.get("supply_chain_tier") and mapped.get("type_description"):
                t = str(mapped["type_description"]).lower()
                tier = 1
                if any(x in t for x in ["refiner", "smelter", "recycler", "processor"]):
                    tier = 2
                elif "trader" in t or "distributor" in t:
                    tier = 3
                mapped["supply_chain_tier"] = tier

            if not mapped.get("industry_sector"):
                src_text = " ".join(filter(None, [
                    str(mapped.get("products_raw") or ""),
                    str(mapped.get("type_description") or ""),
                    str(mapped.get("company_description") or ""),
                ])).lower()
                SECTOR_KW = [
                    ("recycl", "recycled aggregates"),
                    ("rare earth", "metals mining"), ("ree", "metals mining"),
                    ("lithium", "metals mining"), ("cobalt", "metals mining"),
                    ("nickel", "metals mining"), ("copper", "metals mining"),
                    ("zinc", "metals mining"), ("lead", "metals mining"),
                    ("tin", "metals mining"), ("aluminum", "metals mining"),
                    ("aluminium", "metals mining"), ("bauxite", "construction minerals"),
                    ("gold", "metals mining"), ("silver", "metals mining"),
                    ("platinum", "metals mining"), ("uranium", "metals mining"),
                    ("graphite", "industrial minerals"), ("silica", "industrial minerals"),
                    ("potash", "industrial minerals"), ("salt", "industrial minerals"),
                    ("coal", "coal"), ("oil", "oil and gas"), ("gas", "oil and gas"),
                ]
                sector = "metals mining"
                for kw, sec in SECTOR_KW:
                    if kw in src_text:
                        sector = sec
                        break
                mapped["industry_sector"] = sector

            for field_def in schema_fields:
                if "fixed_value" in field_def and field_def["name"] not in mapped:
                    mapped[field_def["name"]] = field_def["fixed_value"]
            mapped["is_verified"] = False

        is_valid, errors = validate_record(mapped, schema_fields)
        if is_valid:
            valid_count += 1
        record = ExtractedRecord(
            job_id=job.id, schema_version=(schema_ver.version if schema_ver else None),
            extraction_confidence=ExtractionConfidence.HIGH if is_valid else ExtractionConfidence.FLAGGED,
            is_schema_valid=is_valid, validation_errors=errors,
            review_status=ReviewStatus.PENDING,
            extracted_fields=mapped, raw_text=raw_text,
            canonical_name=str(mapped.get("canonical_name") or mapped.get("company_name") or mapped.get("material_name") or mapped.get("name") or "")[:512] or None,
        )
        db.add(record)

    invalid_count = len(rows) - valid_count
    source.status = SourceStatus.NEEDS_FIXES if invalid_count > 0 else SourceStatus.READY_FOR_REVIEW
    if not source.extraction_started_at:
        source.extraction_started_at = datetime.now(timezone.utc)
    if invalid_count == 0:
        source.extraction_completed_at = datetime.now(timezone.utc)

    try:
        db.flush()
        _recompute_counts(source, db)
        db.add(AuditLog(
            user_id=current_user.id, project_id=source.project_id,
            action=AuditAction.SOURCE_DATA_UPLOADED,
            after_value={"file": display_filename, "method": extraction_method, "rows": len(rows), "valid": valid_count, "invalid": invalid_count},
        ))
        db.commit()
    except Exception as db_err:
        db.rollback()
        import traceback
        raise HTTPException(
            status_code=500,
            detail=f"DB error saving records: {type(db_err).__name__}: {str(db_err)[:500]}. "
                   f"This usually means a database migration hasn't run. "
                   f"Check /health/db for missing columns. Traceback: {traceback.format_exc()[-800:]}"
        )

    return SourceUploadSummary(
        total_rows=len(rows), valid_rows=valid_count,
        invalid_rows=invalid_count, job_id=job.id,
        extraction_method=extraction_method,
        files_processed=files_processed,
        file_breakdown=file_breakdown,
    )


@router.post("/{source_id}/upload-multi", response_model=SourceUploadSummary)
async def upload_multi_to_source(
    source_id: str,
    request: Request,
    files: list[UploadFile] = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Folder / multi-file upload — for SOPs where source data ships as a folder
    of files rather than one archive.

    Only files classified as "primary_data" by _classify_batch_file become
    ExtractedRecords. Recognised non-data files (manifest/checklist/review
    log) and genuinely unrecognised files are both skipped — never parsed
    into records — with the reason surfaced in each file's breakdown entry.
    This is what stops a folder upload from silently ingesting an unrelated
    file (an index, a stray note, an old backup) as bogus data.
    """
    source = _get_source_or_404(source_id, db)
    if not _is_assigned_extractor(current_user, source):
        raise HTTPException(status_code=403, detail="Only the assigned extractor, project admin, or org admin can upload to this source")

    if not files:
        raise HTTPException(status_code=422, detail="No files provided.")

    content_length = request.headers.get("content-length")
    if content_length and int(content_length) > MAX_UPLOAD_SIZE_BYTES:
        size_mb = int(content_length) / (1024 * 1024)
        limit_mb = MAX_UPLOAD_SIZE_BYTES / (1024 * 1024)
        raise HTTPException(
            status_code=413,
            detail=f"This folder totals {size_mb:.1f}MB — the limit per upload is {limit_mb:.0f}MB. "
                    f"Upload it in smaller batches (a few files at a time).",
        )

    import os as _os

    schema_ver = None
    if source.schema_id:
        schema_ver = db.query(SchemaVersion).filter(
            SchemaVersion.schema_id == source.schema_id
        ).order_by(SchemaVersion.version.desc()).first()
    schema_fields = []
    if schema_ver and schema_ver.definition:
        schema_fields = schema_ver.definition.get("fields", [])

    all_rows: list[dict] = []
    file_breakdown: list[dict] = []
    ai_used = False
    total_size = 0

    def _is_junk_file(fname: str) -> bool:
        junk_exact = {".ds_store", "thumbs.db", "desktop.ini", ".gitkeep", ".gitignore"}
        if fname.lower() in junk_exact:
            return True
        if fname.startswith("~$"):
            return True
        if fname.startswith("."):
            return True
        return False

    real_files = [f for f in files if not _is_junk_file((f.filename or "").split("/")[-1])]

    if not real_files:
        raise HTTPException(
            status_code=422,
            detail="This folder only contains system/hidden files (like .DS_Store or "
                   "Thumbs.db) — no actual data files to upload.",
        )

    all_paths = [(f.filename or "upload").replace("\\", "/") for f in real_files]
    batch_slugs = _derive_batch_slugs(all_paths)

    skipped_unknown: list[str] = []
    recognized_nondata: list[str] = []
    # Every file gets captured here verbatim, regardless of role or whether
    # it parsed successfully — this is the raw preservation layer, entirely
    # independent of the parsing pipeline above it.
    raw_entries: list[dict] = []

    for f, full_path in zip(real_files, all_paths):
        fname = full_path.split("/")[-1]
        ext = _os.path.splitext(fname)[1].lower()

        role = _classify_batch_file(full_path, batch_slugs)

        if role != "primary_data":
            raw = await f.read()
            total_size += len(raw)
            raw_entries.append({"relative_path": full_path, "is_directory": False, "content": raw})
            if role == "unknown":
                skipped_unknown.append(fname)
                file_breakdown.append({
                    "filename": fname, "rows": 0,
                    "error": "not recognised as part of this batch — skipped, not ingested",
                })
            else:
                recognized_nondata.append(fname)
                file_breakdown.append({
                    "filename": fname, "rows": 0,
                    "role": f"{role} (recognised, not extracted as data)",
                })
            continue

        raw = await f.read()
        total_size += len(raw)
        raw_entries.append({"relative_path": full_path, "is_directory": False, "content": raw})

        if not ext or ext not in ALLOWED_UPLOAD_EXTENSIONS:
            file_breakdown.append({"filename": fname, "rows": 0, "error": f"unsupported type '{ext or 'unknown'}'"})
            continue

        try:
            if ext in AI_EXTRACTION_EXTENSIONS:
                _schema_def = schema_ver.definition if schema_ver else {"flexible_validation": True, "fields": []}
                rows = await _extract_with_llm(raw, ext, _schema_def, source)
                ai_used = True
            elif ext == ".zip":
                rows, inner_breakdown = _parse_zip(raw)
                file_breakdown.extend(inner_breakdown)
            else:
                rows = _parse_rows(raw, ext, fname)
        except Exception as e:
            file_breakdown.append({"filename": fname, "rows": 0, "error": str(e)[:150]})
            continue

        for row in rows:
            if isinstance(row, dict) and "_source_file" not in row:
                row["_source_file"] = full_path

        if ext != ".zip":
            file_breakdown.append({"filename": fname, "rows": len(rows)})
        all_rows.extend(rows)

    if not all_rows:
        skipped = [f for f in file_breakdown if f.get("error")]
        detail = f"No primary-data records found across {len(real_files)} file(s)."
        if recognized_nondata:
            detail += f" {len(recognized_nondata)} recognised as non-data (manifest/checklist/review log)."
        if skipped_unknown:
            detail += f" {len(skipped_unknown)} file(s) not recognised and skipped: {', '.join(skipped_unknown[:5])}."
        raise HTTPException(status_code=422, detail=detail)

    extraction_method = "llm" if ai_used else "structured"
    files_processed = len([f for f in file_breakdown if not f.get("error")])

    return _finalize_upload(
        source=source, source_id=source_id, rows=all_rows, ext=".zip",
        extraction_method=extraction_method, file_breakdown=file_breakdown,
        files_processed=files_processed,
        display_filename=f"{len(real_files)} file{'s' if len(real_files) != 1 else ''} (folder upload)",
        content_len=total_size, schema_ver=schema_ver, schema_fields=schema_fields,
        current_user=current_user, db=db, raw_entries=raw_entries,
    )


async def _extract_with_llm(content: bytes, ext: str, schema_definition: dict, source: Source) -> list[dict]:
    """
    Uses Claude to extract structured records from a PDF or plain-text document.
    Returns a list of dicts matching the schema's field names.
    The model is told exactly what fields to extract, their types, and what
    the extraction instructions say — same instructions human extractors follow.
    """
    import re
    from app.core.config import settings

    if ext == ".pdf":
        import pdfplumber, io as _io
        text_parts = []
        with pdfplumber.open(_io.BytesIO(content)) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    text_parts.append(t)
        doc_text = "\n\n".join(text_parts)
    else:
        doc_text = content.decode("utf-8", errors="replace")

    if not doc_text.strip():
        raise ValueError("Could not extract any readable text from the document.")

    MAX_CHARS = 60000
    if len(doc_text) > MAX_CHARS:
        doc_text = doc_text[:MAX_CHARS] + "\n\n[Document truncated — extract from the above portion only]"

    fields = schema_definition.get("fields", [])
    required_fields = [f for f in fields if f.get("required") and "fixed_value" not in f]
    optional_fields = [f for f in fields if not f.get("required") and "fixed_value" not in f]
    fixed_fields = {f["name"]: f["fixed_value"] for f in fields if "fixed_value" in f}

    def field_line(f: dict) -> str:
        parts = [f"- {f['name']} ({f.get('type','string')})"]
        if f.get("description"):
            parts.append(f": {f['description']}")
        if f.get("enum"):
            parts.append(f" — allowed values: {f['enum']}")
        return "".join(parts)

    extraction_instructions = schema_definition.get("extraction_instructions", "")
    grouping_key = schema_definition.get("grouping_key", "")

    system_prompt = f"""You are a precise data extraction specialist. You extract structured records from documents and return them as JSON.

SCHEMA: {schema_definition.get('name', 'Data Extraction Schema')}
{f'GROUPING: Each record represents one unique {grouping_key}.' if grouping_key else ''}
{f'EXTRACTION RULES:\\n{extraction_instructions}' if extraction_instructions else ''}

REQUIRED FIELDS (must be present in every record):
{chr(10).join(field_line(f) for f in required_fields) or '(none)'}

OPTIONAL FIELDS (include if present in the source):
{chr(10).join(field_line(f) for f in optional_fields) or '(none)'}

FIXED FIELDS (always set these exact values, do not extract from document):
{chr(10).join(f'- {k}: {v}' for k, v in fixed_fields.items()) or '(none)'}

CRITICAL — NESTED OBJECT STRUCTURE:
Array fields must contain OBJECTS, not plain strings. Use these exact structures:

manufacturing_sites must be an array of objects:
  {{"location": "Site Name (Grid Ref or Lat/Long)", "country": "Country name", "site_type": "mine|quarry|pit|refinery|smelter|processing plant|handling site|wharf|recycling facility|peat workings|exploration site|laboratory", "raw": "verbatim source text about this site — include ownership %, production figures, URLs"}}

products_offered must be an array of objects:
  {{"product_name": "Product Name", "grade": "Grade or variant e.g. Battery-grade", "product_id": "SITE_PRODUCT_GRADE", "category": "COMMODITY CATEGORY", "source_url": "{source.website_url or ''}", "datasheet_url": null, "cross_graph_material_id": null}}

sources must be an array of objects:
  {{"source_name": "Publication or page name", "source_url": "https://...", "doi": null, "tier": "tier1|tier2|tier3"}}

data_completeness_flags must be this exact object (never null):
  {{"review_score": "manual_only", "defect_rate_ppm": "manual_only", "on_time_delivery_rate": "manual_only", "pricing": "api_only", "inventory_levels": "api_only"}}

extras must be an array with ONE object containing any data that doesn't fit above fields:
  [{{"office_address_1": "...", "office_address_2": "...", "any_other_key": "value"}}]
  Use this for: multiple office addresses, contact details per region, licence numbers,
  JV ownership details, regulatory references — anything structured but not fitting BGS fields.

jv_stakes (if present) must be an array of objects:
  {{"site_name": "Site name", "ownership_pct": 44.0, "jv_partners": ["Partner Name"], "country": "Country", "commodity": "Commodity"}}

annual_production (if present) must be an array of objects:
  {{"commodity": "Copper", "volume": "1,058,100", "unit": "tonnes", "year": "2024", "notes": "own sourced"}}

RULES:
1. Extract EVERY record you can find — do not skip any.
2. If a field is not in the document, set it to null or [] (not a plain string).
3. Return ONLY a JSON array. No preamble, no explanation, no markdown code fences.
4. Each element of the array is one record with exactly the field names above.
5. NEVER put plain strings inside manufacturing_sites, products_offered, sources, or extras arrays.

Example of CORRECT manufacturing_sites:
"manufacturing_sites": [{{"location": "Kamoa-Kakula Copper Complex", "country": "Democratic Republic of Congo", "site_type": "mine", "raw": "Kamoa-Kakula, DRC. Ivanhoe 39.6% | Zijin 39.6% | DRC Govt 20%. Largest undeveloped high-grade copper deposit in the world."}}]

Example of WRONG manufacturing_sites (never do this):
"manufacturing_sites": ["Kamoa-Kakula, Democratic Republic of Congo"]"""

    user_message = f"""Extract all records from this document:

---
{doc_text}
---

Return a JSON array of all records found."""

    from google import genai as _genai
    _client = _genai.Client(api_key=settings.GEMINI_API_KEY)
    _response = _client.models.generate_content(
        model=settings.LLM_MODEL,
        contents=system_prompt + "\n\n" + user_message,
    )
    raw_text = _response.text if _response.text else ""

    clean = raw_text.strip()
    if clean.startswith("```"):
        clean = re.sub(r"^```(?:json)?\s*", "", clean)
        clean = re.sub(r"\s*```$", "", clean)

    records = json.loads(clean.strip())
    if not isinstance(records, list):
        raise ValueError(f"Expected a JSON array, got {type(records).__name__}")

    for rec in records:
        for k, v in fixed_fields.items():
            rec[k] = v
        if "_raw_text" not in rec:
            rec["_raw_text"] = ""

    return records


def _parse_rows(content: bytes, ext: str, filename: str) -> list[dict]:
    if ext == ".csv":
        df = pd.read_csv(io.BytesIO(content))
        return df.where(pd.notnull(df), None).to_dict("records")
    if ext in (".xlsx", ".xls"):
        df = pd.read_excel(io.BytesIO(content))
        return df.where(pd.notnull(df), None).to_dict("records")
    if ext == ".json":
        data = json.loads(content.decode("utf-8"))
        if isinstance(data, list):
            return data
        for key in ("items", "records", "data", "rows", "suppliers", "materials"):
            if isinstance(data, dict) and key in data and isinstance(data[key], list):
                return data[key]
        if isinstance(data, dict):
            return [data]
    return []


def _read_raw_zip_entries(content: bytes) -> list[dict]:
    """
    Reads every entry in a ZIP archive verbatim — every file's bytes, and
    every empty-directory entry — with zero filtering, classification, or
    interpretation. Used to persist the exact original upload structure,
    independent of what gets parsed into records. A directory entry (a
    name ending in "/") is captured with content=None; this is the only
    way an empty folder can survive an upload at all, since a browser's
    folder-picker input never reports empty folders to the server.
    """
    import zipfile as zf_mod
    entries: list[dict] = []
    with zf_mod.ZipFile(io.BytesIO(content)) as zf:
        for name in zf.namelist():
            if name.endswith("/"):
                entries.append({"relative_path": name, "is_directory": True, "content": None})
            else:
                raw = zf.read(name)
                entries.append({"relative_path": name, "is_directory": False, "content": raw})
    return entries


def _parse_zip(content: bytes) -> tuple[list[dict], list[dict]]:
    """
    Parse a ZIP archive containing JSON files (or CSV/Excel files).
    Returns (rows, file_breakdown) where file_breakdown is a list of
    {filename, rows, skipped_reason} dicts for the UI summary.

    Applies the same _classify_batch_file role classification as folder
    uploads, so a zipped SOP batch gets identical primary-data-only
    treatment — a manifest/checklist/review-log file inside a ZIP never
    becomes bogus records either, and neither does an unrecognised file.

    Each returned row is tagged with "_source_file" = its full path inside
    the archive (e.g. "NETL_METALLIC/subfolder/id_72.json"), so the UI can
    reconstruct and display the original folder/subfolder structure.
    """
    import zipfile as zf_mod
    import os as _os

    all_rows: list[dict] = []
    breakdown: list[dict] = []
    SUPPORTED = {".json", ".csv", ".xlsx", ".xls"}
    SKIP_PREFIXES = ("__MACOSX", ".", "_", "~$")

    def is_skippable(name: str) -> bool:
        parts = name.split("/")
        return any(p.startswith(SKIP_PREFIXES) for p in parts if p)

    def parse_member(fname: str, data: bytes) -> tuple[list[dict], str | None]:
        ext = _os.path.splitext(fname)[1].lower()
        if ext not in SUPPORTED:
            return [], f"unsupported type {ext}"
        try:
            rows = _parse_rows(data, ext, fname)
            return rows, None
        except Exception as e:
            return [], str(e)[:120]

    def _tag(rows: list[dict], full_path: str) -> list[dict]:
        for row in rows:
            if isinstance(row, dict) and "_source_file" not in row:
                row["_source_file"] = full_path
        return rows

    with zf_mod.ZipFile(io.BytesIO(content)) as zf:
        names = [n for n in zf.namelist() if not is_skippable(n) and not n.endswith("/")]
        batch_slugs = _derive_batch_slugs(names)

        for name in names:
            fname = name.split("/")[-1]
            ext = _os.path.splitext(fname)[1].lower()
            role = _classify_batch_file(name, batch_slugs)

            if role != "primary_data":
                if role == "unknown":
                    breakdown.append({"filename": fname, "rows": 0,
                        "error": "not recognised as part of this batch — skipped, not ingested"})
                else:
                    breakdown.append({"filename": fname, "rows": 0,
                        "role": f"{role} (recognised, not extracted as data)"})
                continue

            if ext == ".zip":
                inner_content = zf.read(name)
                try:
                    with zf_mod.ZipFile(io.BytesIO(inner_content)) as inner_zf:
                        for inner_name in inner_zf.namelist():
                            if is_skippable(inner_name) or inner_name.endswith("/"):
                                continue
                            inner_fname = inner_name.split("/")[-1]
                            rows, err = parse_member(inner_fname, inner_zf.read(inner_name))
                            full_path = f"{name.rsplit('.', 1)[0]}/{inner_name}"
                            rows = _tag(rows, full_path)
                            breakdown.append({"filename": f"{fname}/{inner_fname}", "rows": len(rows), "error": err})
                            all_rows.extend(rows)
                except Exception as e:
                    breakdown.append({"filename": fname, "rows": 0, "error": f"inner ZIP error: {str(e)[:80]}"})
            elif ext in SUPPORTED:
                rows, err = parse_member(fname, zf.read(name))
                rows = _tag(rows, name)
                breakdown.append({"filename": fname, "rows": len(rows), "error": err})
                all_rows.extend(rows)

    return all_rows, breakdown


# ─── Records (fix + review) ──────────────────────────────────────────────────

@router.get("/{source_id}/records", response_model=PaginatedResponse)
def list_source_records(
    source_id: str,
    validity: str = Query(None, description="valid | invalid"),
    review_status: str = Query(None),
    page: int = Query(1, ge=1), page_size: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user),
):
    source = _get_source_or_404(source_id, db)

    user_roles = {r.role.value for r in current_user.roles}
    if "org_admin" not in user_roles:
        member = db.query(ProjectMember).filter(
            ProjectMember.project_id == source.project_id,
            ProjectMember.user_id == current_user.id,
        ).first()
        if not member:
            raise HTTPException(status_code=403, detail="Access denied")

    latest_job = (
        db.query(ExtractionJob)
        .filter(ExtractionJob.source_id == source_id)
        .order_by(ExtractionJob.created_at.desc())
        .first()
    )
    q = db.query(ExtractedRecord).filter(
        ExtractedRecord.job_id == (latest_job.id if latest_job else None)
    )
    if validity == "valid":
        q = q.filter(ExtractedRecord.is_schema_valid == True)
    elif validity == "invalid":
        q = q.filter(ExtractedRecord.is_schema_valid == False)
    if review_status:
        try:
            q = q.filter(ExtractedRecord.review_status == ReviewStatus(review_status))
        except ValueError:
            pass

    total = q.count()
    records = q.order_by(ExtractedRecord.created_at).offset((page - 1) * page_size).limit(page_size).all()
    return PaginatedResponse(
        items=[_serialize_record(r) for r in records],
        total=total, page=page, page_size=page_size, pages=math.ceil(total / page_size) or 1,
    )


@router.patch("/{source_id}/records/{record_id}", response_model=RecordOut)
def fix_record(
    source_id: str, record_id: str, payload: SourceRecordFix,
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user),
):
    source = _get_source_or_404(source_id, db)
    if not _is_assigned_extractor(current_user, source):
        raise HTTPException(status_code=403, detail="Only the assigned extractor, project admin, or org admin can fix records")

    record = db.query(ExtractedRecord).join(ExtractionJob).filter(
        ExtractedRecord.id == record_id, ExtractionJob.source_id == source_id
    ).first()
    if not record:
        raise HTTPException(status_code=404, detail="Record not found in this source")

    schema_ver = db.query(SchemaVersion).filter(
        SchemaVersion.schema_id == source.schema_id
    ).order_by(SchemaVersion.version.desc()).first()
    schema_fields = schema_ver.definition.get("fields", []) if schema_ver else []

    before = dict(record.extracted_fields or {})
    record.extracted_fields = {**(record.extracted_fields or {}), **payload.extracted_fields}
    is_valid, errors = validate_record(record.extracted_fields, schema_fields)
    record.is_schema_valid = is_valid
    record.validation_errors = errors
    record.extraction_confidence = ExtractionConfidence.HIGH if is_valid else ExtractionConfidence.FLAGGED
    if record.review_status == ReviewStatus.REJECTED and is_valid:
        record.review_status = ReviewStatus.PENDING

    db.flush()
    _recompute_counts(source, db)
    if source.invalid_records == 0 and source.status == SourceStatus.NEEDS_FIXES:
        source.status = SourceStatus.READY_FOR_REVIEW
        source.extraction_completed_at = datetime.now(timezone.utc)

    db.add(AuditLog(
        user_id=current_user.id, project_id=source.project_id,
        action=AuditAction.SOURCE_RECORD_FIXED,
        before_value={"fields": before}, after_value={"fields": record.extracted_fields},
    ))
    db.commit()
    db.refresh(record)
    return _serialize_record(record)


@router.post("/{source_id}/records/{record_id}/review", response_model=RecordOut)
def review_source_record(
    source_id: str, record_id: str, payload: SourceRecordReview,
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user),
):
    source = _get_source_or_404(source_id, db)

    if not _is_assigned_reviewer(current_user, source):
        raise HTTPException(status_code=403, detail="Only reviewers, QA leads, project admins, or org admins can review records")

    record = (
        db.query(ExtractedRecord)
        .join(ExtractionJob, ExtractedRecord.job_id == ExtractionJob.id)
        .filter(ExtractedRecord.id == record_id, ExtractionJob.source_id == source_id)
        .first()
    )
    if not record:
        raise HTTPException(status_code=404, detail="Record not found in this source")

    if payload.action not in ("approve", "reject"):
        raise HTTPException(status_code=422, detail="action must be 'approve' or 'reject'")

    now = datetime.now(timezone.utc)

    if record.review_started_at is None:
        record.review_started_at = now

    if payload.action == "approve":
        record.review_status = ReviewStatus.PENDING_ADMIN_REVIEW
        record.revision_count = (record.revision_count or 0) + 1
        record.admin_review_started_at = now
        if payload.note:
            fc = record.reviewer_field_comments or {}
            fc.setdefault("_general", []).append({
                "comment": payload.note, "user": current_user.email,
                "role": "reviewer", "ts": now.isoformat(),
            })
            record.reviewer_field_comments = fc
        db.add(AuditLog(
            user_id=current_user.id, project_id=source.project_id, record_id=record.id,
            action=AuditAction.RECORD_SENT_TO_ADMIN,
            after_value={"note": payload.note, "revision_count": record.revision_count},
        ))
    else:
        record.review_status = ReviewStatus.REJECTED
        record.correction_count = (record.correction_count or 0) + 1
        if payload.note:
            fc = record.reviewer_field_comments or {}
            fc.setdefault("_general", []).append({
                "comment": payload.note, "user": current_user.email,
                "role": "reviewer", "type": "rejection", "ts": now.isoformat(),
            })
            record.reviewer_field_comments = fc
        db.add(AuditLog(
            user_id=current_user.id, project_id=source.project_id, record_id=record.id,
            action=AuditAction.RECORD_REJECTED,
            after_value={"note": payload.note, "correction_count": record.correction_count},
        ))

    record.review_note = payload.note
    record.reviewed_by = current_user.id
    record.reviewed_at = now

    if source.status not in (SourceStatus.IN_REVIEW,):
        source.status = SourceStatus.IN_REVIEW
        source.review_started_at = source.review_started_at or now

    if payload.action == "reject":
        source.status = SourceStatus.CHANGES_REQUESTED
        if source.assigned_extractor_id:
            db.add(Notification(
                user_id=source.assigned_extractor_id,
                title=f"Record sent back in '{source.name}'",
                body=payload.note or "A reviewer sent a record back for fixes.",
                link=f"/sources/{source.id}",
            ))

    db.flush()
    _recompute_counts(source, db)
    db.commit()
    db.refresh(record)
    return _serialize_record(record)


@router.post("/{source_id}/records/{record_id}/admin-review", response_model=RecordOut)
def admin_review_record(
    source_id: str, record_id: str, payload: dict,
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user),
):
    """
    Admin final sign-off on a record that a reviewer has already approved.
      action = "approve" -> record becomes fully APPROVED (delivered)
      action = "return"  -> record goes back to PENDING, correction_count += 1
    """
    source = _get_source_or_404(source_id, db)
    user_roles = {r.role.value for r in current_user.roles}
    if "org_admin" not in user_roles and "project_admin" not in user_roles and "qa_lead" not in user_roles:
        raise HTTPException(status_code=403, detail="Only admins can do the final review")

    record = (
        db.query(ExtractedRecord)
        .join(ExtractionJob, ExtractedRecord.job_id == ExtractionJob.id)
        .filter(ExtractedRecord.id == record_id, ExtractionJob.source_id == source_id)
        .first()
    )
    if not record:
        raise HTTPException(status_code=404, detail="Record not found in this source")

    action = payload.get("action", "approve")
    note = payload.get("note", "")
    field_comments = payload.get("field_comments", {}) or {}
    now = datetime.now(timezone.utc)

    if action == "approve":
        record.review_status = ReviewStatus.APPROVED
        record.admin_reviewed_by = current_user.id
        record.admin_reviewed_at = now
        record.admin_review_note = note
        if field_comments:
            fc = record.reviewer_field_comments or {}
            for field, comment in field_comments.items():
                fc.setdefault(field, []).append({
                    "comment": comment, "user": current_user.email,
                    "role": "admin", "ts": now.isoformat(),
                })
            record.reviewer_field_comments = fc
        db.add(AuditLog(
            user_id=current_user.id, project_id=source.project_id, record_id=record.id,
            action=AuditAction.RECORD_ADMIN_REVIEWED,
            after_value={"action": "approved", "note": note},
        ))
    else:
        record.review_status = ReviewStatus.PENDING
        record.correction_count = (record.correction_count or 0) + 1
        record.admin_review_note = note
        record.extraction_started_at = record.extraction_started_at or now
        if field_comments:
            fc = record.reviewer_field_comments or {}
            for field, comment in field_comments.items():
                fc.setdefault(field, []).append({
                    "comment": comment, "user": current_user.email,
                    "role": "admin", "ts": now.isoformat(), "type": "correction",
                })
            record.reviewer_field_comments = fc
        db.add(AuditLog(
            user_id=current_user.id, project_id=source.project_id, record_id=record.id,
            action=AuditAction.RECORD_RETURNED_FOR_CORRECTION,
            after_value={"note": note, "correction_count": record.correction_count},
        ))

    db.flush()
    _recompute_counts(source, db)
    db.commit()
    db.refresh(record)
    return _serialize_record(record)


@router.get("/{source_id}/records/{record_id}/timeline")
def record_timeline(
    source_id: str, record_id: str,
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user),
):
    """Full audit history for a record, with time-between-steps calculated."""
    source = _get_source_or_404(source_id, db)
    record = (
        db.query(ExtractedRecord)
        .join(ExtractionJob, ExtractedRecord.job_id == ExtractionJob.id)
        .filter(ExtractedRecord.id == record_id, ExtractionJob.source_id == source_id)
        .first()
    )
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")

    logs = (
        db.query(AuditLog)
        .filter(AuditLog.record_id == record_id)
        .order_by(AuditLog.timestamp.asc())
        .all()
    )

    events = []
    prev_ts = record.created_at
    for log in logs:
        delta_secs = None
        if prev_ts and log.timestamp:
            delta_secs = int((log.timestamp - prev_ts).total_seconds())
        events.append({
            "action": log.action.value,
            "user_id": log.user_id,
            "timestamp": log.timestamp.isoformat() if log.timestamp else None,
            "seconds_since_previous": delta_secs,
            "after_value": log.after_value,
        })
        prev_ts = log.timestamp

    return {
        "record_id": record_id,
        "current_status": record.review_status.value,
        "revision_count": record.revision_count or 0,
        "correction_count": record.correction_count or 0,
        "created_at": record.created_at.isoformat() if record.created_at else None,
        "reviewed_at": record.reviewed_at.isoformat() if record.reviewed_at else None,
        "admin_reviewed_at": record.admin_reviewed_at.isoformat() if record.admin_reviewed_at else None,
        "reviewer_field_comments": record.reviewer_field_comments or {},
        "events": events,
    }


@router.post("/{source_id}/approve", response_model=SourceOut)
def approve_source(
    source_id: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_user),
):
    """Mark a source as fully approved. Warns if pending records remain but allows admins to override."""
    source = _get_source_or_404(source_id, db)

    if not _is_assigned_reviewer(current_user, source):
        raise HTTPException(status_code=403, detail="Only reviewers, QA leads, project admins, or org admins can approve a source")

    user_roles = _user_roles(current_user)
    is_admin = "org_admin" in user_roles or "project_admin" in user_roles

    pending_or_rejected = (
        db.query(ExtractedRecord)
        .join(ExtractionJob, ExtractedRecord.job_id == ExtractionJob.id)
        .filter(
            ExtractionJob.source_id == source_id,
            ExtractedRecord.review_status != ReviewStatus.APPROVED,
        ).count()
    )

    if pending_or_rejected > 0 and not is_admin:
        raise HTTPException(
            status_code=422,
            detail=f"{pending_or_rejected} record(s) not yet approved. Approve them first or ask an admin to override."
        )

    source.status = SourceStatus.APPROVED
    source.approved_at = datetime.now(timezone.utc)
    source.review_completed_at = source.review_completed_at or datetime.now(timezone.utc)

    approved_count = (
        db.query(ExtractedRecord)
        .join(ExtractionJob, ExtractedRecord.job_id == ExtractionJob.id)
        .filter(
            ExtractionJob.source_id == source_id,
            ExtractedRecord.review_status == ReviewStatus.APPROVED,
        ).count()
    )
    for job in db.query(ExtractionJob).filter(ExtractionJob.source_id == source_id).all():
        job.status = JobStatus.VALIDATED
        job.total_approved = approved_count

    db.add(AuditLog(
        user_id=current_user.id, project_id=source.project_id,
        action=AuditAction.SOURCE_APPROVED, after_value={"source_id": source_id},
    ))
    db.commit()
    db.refresh(source)
    return _serialize_source(source)


# ─── Export package ──────────────────────────────────────────────────────────

@router.get("/{source_id}/files")
def list_uploaded_files(
    source_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Returns the exact raw structure of the most recent upload for this
    source — every file and (for ZIP uploads) every empty directory, with
    its path and size, independent of what was parsed into records. This
    is the literal folder/ZIP contents as uploaded, not the extracted data.
    """
    source = _get_source_or_404(source_id, db)
    if not _can_access(current_user, source.project):
        raise HTTPException(status_code=403, detail="Access denied")

    latest_job = (
        db.query(ExtractionJob)
        .filter(ExtractionJob.source_id == source_id)
        .order_by(ExtractionJob.created_at.desc())
        .first()
    )
    if not latest_job:
        return {"job_id": None, "entries": [], "total_files": 0, "total_directories": 0}

    entries = (
        db.query(UploadedFileEntry)
        .filter(UploadedFileEntry.job_id == latest_job.id)
        .order_by(UploadedFileEntry.relative_path)
        .all()
    )

    return {
        "job_id": latest_job.id,
        "entries": [
            {
                "id": e.id,
                "relative_path": e.relative_path,
                "is_directory": e.is_directory,
                "size_bytes": e.size_bytes or 0,
                "review_status": e.review_status,
                "review_note": e.review_note,
                "reviewed_at": e.reviewed_at.isoformat() if e.reviewed_at else None,
            }
            for e in entries
        ],
        "total_files": sum(1 for e in entries if not e.is_directory),
        "total_directories": sum(1 for e in entries if e.is_directory),
    }


def _get_file_entry_for_source(source_id: str, file_id: str, db: Session) -> "UploadedFileEntry":
    """
    Looks up one raw file entry by its own id, and confirms it actually
    belongs to this source (via its job's source_id) — so a file id from a
    different source can never be previewed or reviewed through this one.
    """
    entry = db.query(UploadedFileEntry).filter(UploadedFileEntry.id == file_id).first()
    if not entry:
        raise HTTPException(status_code=404, detail="File not found")
    job = db.query(ExtractionJob).filter(ExtractionJob.id == entry.job_id).first()
    if not job or job.source_id != source_id:
        raise HTTPException(status_code=404, detail="File not found for this source")
    return entry


@router.get("/{source_id}/files/{file_id}/content")
def get_uploaded_file_content(
    source_id: str,
    file_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Returns the decoded text content of one raw uploaded file, for previewing
    inline — JSON/CSV/plain-text files decode cleanly and are shown as-is
    (JSON gets pretty-printed client-side). Binary or non-UTF-8 content
    returns a clear "can't preview this" response instead of garbled bytes,
    with Download Original as the fallback.
    """
    source = _get_source_or_404(source_id, db)
    if not _can_access(current_user, source.project):
        raise HTTPException(status_code=403, detail="Access denied")

    entry = _get_file_entry_for_source(source_id, file_id, db)
    if entry.is_directory:
        raise HTTPException(status_code=422, detail="This entry is a folder, not a file.")

    import os as _os
    ext = _os.path.splitext(entry.relative_path)[1].lower()

    raw = entry.content or b""
    try:
        text_content = raw.decode("utf-8")
    except UnicodeDecodeError:
        return {
            "id": entry.id,
            "relative_path": entry.relative_path,
            "kind": "unsupported",
            "content": None,
            "review_status": entry.review_status,
            "review_note": entry.review_note,
            "reviewed_at": entry.reviewed_at.isoformat() if entry.reviewed_at else None,
        }

    kind = "json" if ext == ".json" else "csv" if ext == ".csv" else "text"

    return {
        "id": entry.id,
        "relative_path": entry.relative_path,
        "kind": kind,
        "content": text_content,
        "review_status": entry.review_status,
        "review_note": entry.review_note,
        "reviewed_at": entry.reviewed_at.isoformat() if entry.reviewed_at else None,
    }


class FileReviewRequest(BaseModel):
    action: str  # "approve" | "reject"
    note: str = ""


@router.post("/{source_id}/files/{file_id}/review")
def review_uploaded_file(
    source_id: str,
    file_id: str,
    payload: FileReviewRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Approve or reject one raw uploaded file — a manifest, QA checklist,
    review log, or any other supporting file. A lightweight parallel to
    record review, for files that never become ExtractedRecords and so
    previously had no review step of their own at all.
    """
    source = _get_source_or_404(source_id, db)
    if not _is_assigned_reviewer(current_user, source):
        raise HTTPException(status_code=403, detail="Only reviewers, QA leads, project admins, or org admins can review files")

    if payload.action not in ("approve", "reject"):
        raise HTTPException(status_code=422, detail="action must be 'approve' or 'reject'")

    entry = _get_file_entry_for_source(source_id, file_id, db)
    if entry.is_directory:
        raise HTTPException(status_code=422, detail="Folders can't be reviewed.")

    entry.review_status = "approved" if payload.action == "approve" else "rejected"
    entry.review_note = payload.note
    entry.reviewed_by = current_user.id
    entry.reviewed_at = datetime.now(timezone.utc)
    db.commit()

    return {
        "id": entry.id,
        "relative_path": entry.relative_path,
        "review_status": entry.review_status,
        "review_note": entry.review_note,
        "reviewed_at": entry.reviewed_at.isoformat(),
    }



@router.get("/{source_id}/files/download")
def download_uploaded_files(
    source_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Reconstructs and returns a ZIP built exactly from the stored raw
    entries — every file's original bytes, and every empty directory
    entry — so downloading this gives back precisely what was uploaded,
    the same way downloading a ZIP from Google Drive returns exactly what
    was put in it.
    """
    source = _get_source_or_404(source_id, db)
    if not _can_access(current_user, source.project):
        raise HTTPException(status_code=403, detail="Access denied")

    latest_job = (
        db.query(ExtractionJob)
        .filter(ExtractionJob.source_id == source_id)
        .order_by(ExtractionJob.created_at.desc())
        .first()
    )
    if not latest_job:
        raise HTTPException(status_code=404, detail="No upload found for this source.")

    entries = (
        db.query(UploadedFileEntry)
        .filter(UploadedFileEntry.job_id == latest_job.id)
        .order_by(UploadedFileEntry.relative_path)
        .all()
    )
    if not entries:
        raise HTTPException(status_code=404, detail="No raw files stored for this source's latest upload.")

    import zipfile
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for e in entries:
            if e.is_directory:
                # A directory entry in a ZIP is just a name ending in "/"
                # with no data — this is what lets an empty folder survive
                # a round trip through the archive at all.
                path = e.relative_path if e.relative_path.endswith("/") else e.relative_path + "/"
                zf.writestr(path, b"")
            else:
                zf.writestr(e.relative_path, e.content or b"")

    buf.seek(0)
    cn = source.canonical_name if hasattr(source, "canonical_name") and source.canonical_name else source.name
    cn = str(cn).strip().replace("/", "-").replace("\\", "-").replace(":", "-") or "source"
    return StreamingResponse(
        buf, media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{cn}_original_upload.zip"'},
    )


@router.get("/{source_id}/export")
def export_source(source_id: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    source = _get_source_or_404(source_id, db)
    if not _is_assigned_reviewer(current_user, source) and not _is_assigned_extractor(current_user, source):
        raise HTTPException(status_code=403, detail="Access denied")
    if source.status != SourceStatus.APPROVED:
        raise HTTPException(status_code=422, detail="Source must be approved before export")

    records = (
        db.query(ExtractedRecord)
        .join(ExtractionJob, ExtractedRecord.job_id == ExtractionJob.id)
        .filter(
            ExtractionJob.source_id == source_id,
            ExtractedRecord.review_status == ReviewStatus.APPROVED,
        ).all()
    )
    data = [r.extracted_fields for r in records]

    duration = None
    if source.extraction_started_at and source.approved_at:
        delta = source.approved_at - source.extraction_started_at
        hours = delta.total_seconds() / 3600
        duration = f"{hours:.1f} hours" if hours < 48 else f"{delta.days} days"

    cover_sheet = f"""# {source.name} — Data Export Cover Sheet

## Source Information
- **Source name:** {source.name}
- **Description:** {source.description or '(none)'}
- **Website:** {source.website_url or '(none)'}
- **Schema:** {source.schema.name if source.schema else 'Unknown'}
- **Project:** {source.project.name if source.project else 'Unknown'}

## Extraction Summary
- **Total rows uploaded:** {source.total_records}
- **Passed schema validation:** {source.valid_records}
- **Failed validation (fixed before approval):** {source.invalid_records}
- **Final approved records:** {len(records)}

## Team
- **Extractor:** {source.extractor.full_name if source.extractor else 'Unassigned'}
- **Reviewer:** {source.reviewer.full_name if source.reviewer else 'Unassigned'}

## Timeline
- **Extraction started:** {source.extraction_started_at.strftime('%Y-%m-%d %H:%M UTC') if source.extraction_started_at else 'N/A'}
- **Extraction completed:** {source.extraction_completed_at.strftime('%Y-%m-%d %H:%M UTC') if source.extraction_completed_at else 'N/A'}
- **Review started:** {source.review_started_at.strftime('%Y-%m-%d %H:%M UTC') if source.review_started_at else 'N/A'}
- **Approved:** {source.approved_at.strftime('%Y-%m-%d %H:%M UTC') if source.approved_at else 'N/A'}
- **Total time, start to approval:** {duration or 'N/A'}

## Notes / Assumptions
{source.notes or '(none recorded)'}

---
Generated by Xtrium DataOps on {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}
"""

    import zipfile
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("data.json", json.dumps(data, indent=2, ensure_ascii=False, default=str))
        zf.writestr("COVER_SHEET.md", cover_sheet)

        latest_job = db.query(ExtractionJob).filter(
            ExtractionJob.source_id == source_id
        ).order_by(ExtractionJob.created_at.desc()).first()
        if latest_job and latest_job.source_file_url:
            try:
                from app.services.storage import storage
                raw_bytes = storage.read(latest_job.source_file_url)
                raw_name = latest_job.source_file_name or "raw_upload"
                zf.writestr(f"raw_{raw_name}", raw_bytes)
            except Exception:
                pass

    buf.seek(0)
    cn = source.canonical_name or source.name
    cn = str(cn).strip().replace("/", "-").replace("\\", "-").replace(":", "-") or "source"
    return StreamingResponse(
        buf, media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{cn}.zip"'},
    )


# ─── Client Timesheet Export ─────────────────────────────────────────────────

@router.get("/export/timesheet")
def export_timesheet(
    project_id: str | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Downloadable Excel sheet for clients showing, per source:
      - when extraction started
      - when it was fully delivered (all records admin-approved)
      - how long it took
      - who worked on it (extractor / reviewer / admin)

    A source only gets a "Delivered At" timestamp once every one of its
    records has been through admin final review and approved — matches
    the double-review workflow (reviewer approve -> admin approve).
    """
    q = db.query(Source)
    if project_id:
        q = q.filter(Source.project_id == project_id)
    sources = q.order_by(Source.created_at.asc()).all()

    if not sources:
        raise HTTPException(status_code=404, detail="No sources found for this project")

    rows = []
    correction_rows = []
    activity_rows = []

    def _uname(uid: str | None) -> str:
        if not uid:
            return ""
        u = db.query(User).filter(User.id == uid).first()
        return u.full_name if u else ""

    for s in sources:
        jobs = db.query(ExtractionJob).filter(ExtractionJob.source_id == s.id).all()
        job_ids = [j.id for j in jobs]
        records = (
            db.query(ExtractedRecord).filter(ExtractedRecord.job_id.in_(job_ids)).all()
            if job_ids else []
        )

        for r in records:
            record_label = None
            if isinstance(r.extracted_fields, dict):
                record_label = r.extracted_fields.get("_source_file")
                if record_label:
                    record_label = record_label.split("/")[-1]
            record_label = record_label or (r.canonical_name if getattr(r, "canonical_name", None) else r.id[:8])

            comments = r.reviewer_field_comments or {}
            has_any_comment = any(comments.values()) if isinstance(comments, dict) else False

            if (r.correction_count or 0) > 0 or has_any_comment:
                if isinstance(comments, dict) and comments:
                    for field, entries in comments.items():
                        for entry in (entries or []):
                            correction_rows.append({
                                "Source": s.name,
                                "Record": record_label,
                                "Times Returned": r.correction_count or 0,
                                "Field": "General" if field == "_general" else field,
                                "Feedback": entry.get("comment", ""),
                                "By": entry.get("user", ""),
                                "Role": (entry.get("role", "") or "").title(),
                                "Type": "Correction" if entry.get("type") in ("correction", "rejection") else "Note",
                                "When": entry.get("ts", "")[:16].replace("T", " ") if entry.get("ts") else "",
                            })
                else:
                    correction_rows.append({
                        "Source": s.name, "Record": record_label,
                        "Times Returned": r.correction_count or 0,
                        "Field": "", "Feedback": "(no comment recorded)",
                        "By": "", "Role": "", "Type": "Correction", "When": "",
                    })

        from sqlalchemy import or_ as _or_
        record_ids = [r.id for r in records]
        audit_entries = (
            db.query(AuditLog)
            .filter(_or_(AuditLog.source_id == s.id, AuditLog.record_id.in_(record_ids)))
            .order_by(AuditLog.timestamp.asc())
            .all()
        )
        for e in audit_entries:
            note = ""
            if isinstance(e.after_value, dict):
                note = e.after_value.get("note") or e.after_value.get("reason") or e.after_value.get("action") or ""
            activity_rows.append({
                "Source": s.name,
                "Action": e.action.value.replace("_", " ").title(),
                "By": _uname(e.user_id),
                "Note": note,
                "When": e.timestamp.strftime("%Y-%m-%d %H:%M") if e.timestamp else "",
            })

        total = len(records)
        approved = [r for r in records if r.review_status == ReviewStatus.APPROVED]
        fully_delivered = total > 0 and len(approved) == total

        delivered_at = None
        if fully_delivered:
            admin_times = [r.admin_reviewed_at for r in approved if r.admin_reviewed_at]
            if admin_times:
                delivered_at = max(admin_times)

        started_at = s.extraction_started_at

        duration_str = ""
        if started_at and delivered_at:
            delta = delivered_at - started_at
            total_hours = delta.total_seconds() / 3600
            if total_hours < 24:
                duration_str = f"{total_hours:.1f} hours"
            else:
                duration_str = f"{delta.days} days, {(delta.seconds // 3600)} hours"

        extractor_name = ""
        if getattr(s, "assigned_extractor_id", None):
            u = db.query(User).filter(User.id == s.assigned_extractor_id).first()
            extractor_name = u.full_name if u else ""

        reviewer_name = ""
        if getattr(s, "assigned_reviewer_id", None):
            u = db.query(User).filter(User.id == s.assigned_reviewer_id).first()
            reviewer_name = u.full_name if u else ""

        admin_name = ""
        if approved:
            admin_ids = [r.admin_reviewed_by for r in approved if r.admin_reviewed_by]
            if admin_ids:
                u = db.query(User).filter(User.id == admin_ids[-1]).first()
                admin_name = u.full_name if u else ""

        def _stage_duration(t1, t2):
            if not t1 or not t2:
                return ""
            d = t2 - t1
            h = d.total_seconds() / 3600
            return f"{h:.1f}h" if h < 24 else f"{d.days}d {int(h % 24)}h"

        llm_start = getattr(s, "llm_verification_started_at", None)
        llm_end = getattr(s, "llm_verification_completed_at", None)

        rows.append({
            "Source": s.name,
            "Status": "Delivered" if fully_delivered else s.status.value.replace("_", " ").title(),
            "Extraction Started": started_at.strftime("%Y-%m-%d %H:%M") if started_at else "",
            "Extraction Duration": _stage_duration(started_at, s.extraction_completed_at),
            "LLM Verification Duration": _stage_duration(llm_start, llm_end),
            "Review Duration": _stage_duration(s.review_started_at, s.review_completed_at),
            "Delivered At": delivered_at.strftime("%Y-%m-%d %H:%M") if delivered_at else "",
            "Total Duration": duration_str,
            "Resets": getattr(s, "reset_count", 0) or 0,
            "Total Records": total,
            "Approved Records": len(approved),
            "Extractor": extractor_name,
            "Reviewer": reviewer_name,
            "Admin Approved By": admin_name,
        })

    df = pd.DataFrame(rows)
    df_corrections = pd.DataFrame(correction_rows)
    df_activity = pd.DataFrame(activity_rows)

    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    def _style_sheet(writer, df_, sheet_name, header_color, widths):
        if df_.empty:
            df_ = pd.DataFrame([{"Info": "No data for this section"}])
        df_.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        for col_idx, col_name in enumerate(df_.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col_idx, col_name in enumerate(df_.columns, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_name, 18)
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 22
        for col_idx, col_name in enumerate(df_.columns, 1):
            if col_name in ("Feedback", "Note"):
                for row_idx in range(2, len(df_) + 2):
                    ws.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        _style_sheet(writer, df, "Delivery Timesheet", "2563EB", {
            "Source": 38, "Status": 16, "Extraction Started": 18,
            "Extraction Duration": 18, "LLM Verification Duration": 20, "Review Duration": 16,
            "Delivered At": 18, "Total Duration": 16, "Resets": 10, "Total Records": 14, "Approved Records": 16,
            "Extractor": 20, "Reviewer": 20, "Admin Approved By": 20,
        })
        _style_sheet(writer, df_corrections, "Corrections & Feedback", "DC2626", {
            "Source": 34, "Record": 24, "Times Returned": 14, "Field": 20,
            "Feedback": 50, "By": 20, "Role": 14, "Type": 14, "When": 18,
        })
        _style_sheet(writer, df_activity, "Full Activity Log", "7C3AED", {
            "Source": 34, "Action": 26, "By": 20, "Note": 50, "When": 18,
        })

    buf.seek(0)
    filename = f"delivery_timesheet_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )



# ─── Performance analytics ───────────────────────────────────────────────────

@router.get("/stats/performance")
def performance_stats(
    project_id: str = Query(None),
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user),
):
    user_roles = {r.role.value for r in current_user.roles}
    if not user_roles.intersection({"org_admin", "project_admin", "qa_lead"}):
        raise HTTPException(status_code=403, detail="Admin or QA access required")

    q = db.query(Source)
    if project_id:
        q = q.filter(Source.project_id == project_id)
    sources = q.all()

    by_extractor: dict[str, dict] = {}
    by_reviewer: dict[str, dict] = {}

    for s in sources:
        if s.assigned_extractor_id:
            stats = by_extractor.setdefault(s.assigned_extractor_id, {
                "user_id": s.assigned_extractor_id,
                "name": s.extractor.full_name if s.extractor else "Unknown",
                "sources_count": 0, "approved_count": 0, "total_hours": 0.0, "samples": 0,
            })
            stats["sources_count"] += 1
            if s.status == SourceStatus.APPROVED:
                stats["approved_count"] += 1
            if s.extraction_started_at and s.extraction_completed_at:
                hours = (s.extraction_completed_at - s.extraction_started_at).total_seconds() / 3600
                stats["total_hours"] += hours
                stats["samples"] += 1

        if s.assigned_reviewer_id:
            stats = by_reviewer.setdefault(s.assigned_reviewer_id, {
                "user_id": s.assigned_reviewer_id,
                "name": s.reviewer.full_name if s.reviewer else "Unknown",
                "sources_count": 0, "approved_count": 0, "total_hours": 0.0, "samples": 0,
            })
            stats["sources_count"] += 1
            if s.status == SourceStatus.APPROVED:
                stats["approved_count"] += 1
            if s.review_started_at and s.review_completed_at:
                hours = (s.review_completed_at - s.review_started_at).total_seconds() / 3600
                stats["total_hours"] += hours
                stats["samples"] += 1

    def finalize(d: dict) -> list[dict]:
        out = []
        for v in d.values():
            avg = v["total_hours"] / v["samples"] if v["samples"] else None
            out.append({**v, "avg_hours_per_source": round(avg, 1) if avg is not None else None})
        return out

    return {"extractors": finalize(by_extractor), "reviewers": finalize(by_reviewer)}


# ─── Delete source ────────────────────────────────────────────────────────────

@router.delete("/{source_id}", status_code=204)
def delete_source(
    source_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Delete a source and all its jobs and records. Only admins can do this."""
    source = db.query(Source).filter(Source.id == source_id).first()
    if not source:
        raise HTTPException(status_code=404, detail="Source not found")

    user_roles = {r.role.value for r in current_user.roles}
    is_org_admin = "org_admin" in user_roles

    if not is_org_admin:
        from app.models.all_models import ProjectMember as PM
        member = db.query(PM).filter(
            PM.project_id == source.project_id,
            PM.user_id == current_user.id,
        ).first()
        if not member or member.role.value not in ("project_admin", "org_admin"):
            raise HTTPException(status_code=403, detail="Only admins can delete sources")

    if source.status == SourceStatus.APPROVED and not is_org_admin:
        raise HTTPException(status_code=422, detail="Approved sources cannot be deleted. Reset it first, or ask an org admin to delete it.")

    source_name = source.name
    source_status = source.status.value
    project_id = source.project_id

    job_ids = [j.id for j in db.query(ExtractionJob).filter(ExtractionJob.source_id == source_id).all()]
    if job_ids:
        db.query(ExtractedRecord).filter(ExtractedRecord.job_id.in_(job_ids)).delete(synchronize_session=False)
        db.query(ExtractionJob).filter(ExtractionJob.id.in_(job_ids)).delete(synchronize_session=False)

    db.delete(source)
    db.add(AuditLog(
        user_id=current_user.id, project_id=project_id,
        action=AuditAction.SOURCE_STATUS_CHANGED,
        before_value={"name": source_name, "status": source_status},
        after_value={"deleted": True},
    ))
    db.commit()


@router.delete("/{source_id}/records/{record_id}", status_code=204)
def delete_source_record(
    source_id: str, record_id: str,
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user),
):
    """Delete a single record from a source. Extractor or admin only."""
    source = _get_source_or_404(source_id, db)
    if not _is_assigned_extractor(current_user, source):
        raise HTTPException(status_code=403, detail="Only the assigned extractor or admin can delete records")
    record = db.query(ExtractedRecord).join(ExtractionJob).filter(
        ExtractedRecord.id == record_id, ExtractionJob.source_id == source_id
    ).first()
    if not record:
        raise HTTPException(status_code=404, detail="Record not found in this source")
    db.delete(record)
    db.flush()
    _recompute_counts(source, db)
    db.commit()


# ─── Capability 1: Scrape website → extract records via AI ──────────────────

@router.post("/{source_id}/scrape", response_model=SourceUploadSummary)
async def scrape_source_website(
    source_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Fetches the source's website_url, extracts readable text, then runs the
    same AI extraction pipeline as PDF/TXT uploads. Records are created and
    validated against the schema automatically.

    This is the 'auto-scrape' capability — point the source at a URL and let
    Claude pull the structured records directly.
    """
    source = _get_source_or_404(source_id, db)
    if not _is_assigned_extractor(current_user, source):
        raise HTTPException(status_code=403, detail="Only the assigned extractor or admin can scrape this source")

    if not source.website_url:
        raise HTTPException(status_code=422, detail="This source has no website URL set. Edit the source to add one first.")

    schema_ver = None
    if source.schema_id:
        schema_ver = db.query(SchemaVersion).filter(
            SchemaVersion.schema_id == source.schema_id
        ).order_by(SchemaVersion.version.desc()).first()
        if not schema_ver:
            raise HTTPException(status_code=422, detail="This source's schema has no field definitions yet.")

    from app.services.web_scraper import fetch_url_text
    try:
        web_text, meta = await fetch_url_text(source.website_url)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Could not fetch {source.website_url}: {str(e)}")

    if not web_text.strip():
        raise HTTPException(status_code=422, detail="The page returned no readable text. It may require JavaScript or a login.")

    _schema_def = schema_ver.definition if schema_ver else {"flexible_validation": True, "fields": [],
        "extraction_instructions": "Extract all relevant data from the page matching the task description."}
    schema_fields = _schema_def.get("fields", [])
    try:
        rows = await _extract_with_llm(web_text.encode("utf-8"), ".txt", _schema_def, source)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"AI extraction failed: {str(e)}")

    if not rows:
        raise HTTPException(status_code=422, detail="AI found no records on this page. The page structure may not match what was expected, or the content is behind a login.")

    _safe_clear_old_jobs(source_id, db)

    job = ExtractionJob(
        project_id=source.project_id, source_id=source_id,
        schema_id=source.schema_id, schema_version=(schema_ver.version if schema_ver else None),
        name=f"{source.name} — web scrape {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')}",
        source_file_name=source.website_url[:255], source_file_size_bytes=len(web_text),
        source_type=FileSourceType.CSV, status=JobStatus.READY_FOR_REVIEW,
        total_raw_records=len(rows), total_extracted=len(rows),
        created_by=current_user.id,
    )
    db.add(job)
    db.flush()

    valid_count = 0
    for row in rows:
        mapped = {k: v for k, v in row.items() if k != "_raw_text"}
        is_valid, errors = validate_record(mapped, schema_fields)
        if is_valid:
            valid_count += 1
        record = ExtractedRecord(
            job_id=job.id, schema_version=(schema_ver.version if schema_ver else None),
            extraction_confidence=ExtractionConfidence.HIGH if is_valid else ExtractionConfidence.FLAGGED,
            is_schema_valid=is_valid, validation_errors=errors,
            review_status=ReviewStatus.PENDING,
            extracted_fields=mapped,
            raw_text=f"[Scraped from {source.website_url}]\n\n{web_text[:2000]}",
            canonical_name=str(mapped.get("canonical_name") or mapped.get("company_name") or mapped.get("material_name") or "")[:512] or None,
        )
        db.add(record)

    invalid_count = len(rows) - valid_count
    source.status = SourceStatus.NEEDS_FIXES if invalid_count > 0 else SourceStatus.READY_FOR_REVIEW
    if not source.extraction_started_at:
        source.extraction_started_at = datetime.now(timezone.utc)
    if invalid_count == 0:
        source.extraction_completed_at = datetime.now(timezone.utc)

    db.flush()
    _recompute_counts(source, db)
    db.add(AuditLog(
        user_id=current_user.id, project_id=source.project_id,
        action=AuditAction.SOURCE_DATA_UPLOADED,
        after_value={"method": "web_scrape", "url": source.website_url, "rows": len(rows), "valid": valid_count},
    ))
    db.commit()

    return SourceUploadSummary(
        total_rows=len(rows), valid_rows=valid_count,
        invalid_rows=invalid_count, job_id=job.id,
        extraction_method="llm",
    )


# ─── Capability 2: LLM verification — cross-check records against live website ─

@router.post("/{source_id}/verify")
async def llm_verify_source(
    source_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Runs the LLM verification stage:
      1. Fetches the live source website
      2. Batches all extracted records (20 at a time)
      3. Claude cross-checks each record's field values against the actual page content
      4. Stores per-record web_check_flags with specific field issues and suggested corrections
      5. Returns a summary of what passed, what was flagged, and what to fix
    """
    source = _get_source_or_404(source_id, db)
    if not _is_assigned_reviewer(current_user, source):
        raise HTTPException(status_code=403, detail="Access denied")

    if not source.website_url:
        raise HTTPException(status_code=422, detail="No website URL — cannot verify without a source to check against.")

    schema_ver = db.query(SchemaVersion).filter(
        SchemaVersion.schema_id == source.schema_id
    ).order_by(SchemaVersion.version.desc()).first()

    records = db.query(ExtractedRecord).join(ExtractionJob).filter(
        ExtractionJob.source_id == source_id,
    ).all()

    if not records:
        raise HTTPException(status_code=422, detail="No records to verify.")

    previous_status = source.status
    if not source.llm_verification_started_at:
        source.llm_verification_started_at = datetime.now(timezone.utc)
    source.status = SourceStatus.LLM_VERIFICATION
    db.flush()

    from app.services.web_scraper import fetch_url_text
    try:
        web_text, meta = await fetch_url_text(source.website_url)
    except Exception as e:
        source.status = previous_status
        db.commit()
        raise HTTPException(status_code=502, detail=f"Could not fetch source website: {str(e)}")

    if not web_text.strip():
        source.status = previous_status
        db.commit()
        raise HTTPException(status_code=422, detail="The source website returned no readable text — may require JavaScript or login.")

    schema_definition = schema_ver.definition if schema_ver else {}
    fields_def = schema_definition.get("fields", [])
    field_context = [
        {
            "name": f["name"],
            "type": f.get("type", "string"),
            "required": f.get("required", False),
            "description": f.get("description", ""),
            "enum": f.get("enum", []),
        }
        for f in fields_def if "fixed_value" not in f
    ]
    extraction_instructions = schema_definition.get("extraction_instructions", "")

    from app.core.config import settings
    from google import genai as _genai
    _client = _genai.Client(api_key=settings.GEMINI_API_KEY)

    BATCH_SIZE = 20
    total = len(records)
    verified_count = 0
    flagged_count = 0
    error_count = 0

    system_prompt = f"""You are a data quality verifier for a structured data extraction system.

Your job: cross-check extracted records against the actual text from the source website and flag any discrepancies.

Schema: {schema_definition.get('name', 'Data Schema')}
{f'Extraction rules: {extraction_instructions}' if extraction_instructions else ''}

Schema fields to verify:
{json.dumps(field_context, indent=2)}

The source website text will be provided. For each record, check whether its field values are:
1. Supported by the website content (PASS)
2. Inconsistent or likely wrong (FLAG with specific correction)
3. Unverifiable from this page (SKIP - note it cannot be confirmed)

Respond ONLY with valid JSON. No markdown, no preamble.

Required format:
{{
  "results": [
    {{
      "record_id": "<id>",
      "verdict": "PASS" | "FLAG" | "SKIP",
      "summary": "<one sentence>",
      "flags": [
        {{
          "field": "<field_name>",
          "issue": "<what's wrong>",
          "suggested_value": "<what the website says it should be>",
          "confidence": 0.0-1.0
        }}
      ]
    }}
  ]
}}"""

    for batch_start in range(0, total, BATCH_SIZE):
        batch = records[batch_start: batch_start + BATCH_SIZE]

        records_payload = [
            {"record_id": r.id, "fields": r.extracted_fields}
            for r in batch
        ]

        user_content = json.dumps({
            "website_text": web_text[:50000],
            "records_to_verify": records_payload,
        }, ensure_ascii=False, default=str)

        try:
            _response = _client.models.generate_content(
                model=settings.LLM_MODEL,
                contents=system_prompt + "\n\n" + user_content,
            )
            raw = _response.text if _response.text else ""

            import re as _re
            clean = raw.strip()
            if clean.startswith("```"):
                clean = _re.sub(r"^```(?:json)?\s*", "", clean)
                clean = _re.sub(r"\s*```$", "", clean)

            result = json.loads(clean.strip())
            batch_results = {r["record_id"]: r for r in result.get("results", [])}

        except Exception as e:
            for r in batch:
                r.web_verified = None
                r.web_check_summary = f"Verification error: {str(e)[:200]}"
            error_count += len(batch)
            db.flush()
            continue

        for record in batch:
            rec_result = batch_results.get(record.id, {})
            verdict = rec_result.get("verdict", "SKIP")
            flags = rec_result.get("flags", [])
            summary = rec_result.get("summary", "")

            record.web_verified = (verdict == "PASS")
            record.web_check_flags = flags
            record.web_check_summary = summary

            if verdict == "PASS":
                verified_count += 1
            elif verdict == "FLAG":
                flagged_count += 1
            else:
                error_count += 1

            db.add(AuditLog(
                user_id=current_user.id, project_id=source.project_id, record_id=record.id,
                action=AuditAction.SOURCE_STATUS_CHANGED,
                after_value={"stage": "llm_verification", "verdict": verdict, "summary": summary},
            ))

        db.flush()

    source.llm_verification_completed_at = datetime.now(timezone.utc)
    source.status = SourceStatus.READY_FOR_REVIEW
    db.add(AuditLog(
        user_id=current_user.id, project_id=source.project_id,
        action=AuditAction.SOURCE_STATUS_CHANGED,
        before_value={"status": previous_status.value},
        after_value={"status": "ready_for_review", "stage": "llm_verification_complete",
                     "verified": verified_count, "flagged": flagged_count},
    ))
    db.commit()

    return {
        "total_records": total,
        "verified": verified_count,
        "flagged": flagged_count,
        "unverifiable": error_count,
        "website_url": source.website_url,
        "website_chars_read": meta.get("char_count", 0),
        "truncated": meta.get("truncated", False),
        "message": (
            f"Verification complete — {verified_count} records match the source website, "
            f"{flagged_count} have field-level issues to review."
        ),
    }


# ─── Capability 3: Schema definition endpoint (for review panel) ─────────────

@router.get("/{source_id}/schema")
def get_source_schema(
    source_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Returns the full schema definition for this source — used by the review UI
    to show field descriptions, types, and allowed values alongside each record.
    """
    source = _get_source_or_404(source_id, db)
    if not _can_access(current_user, source.project):
        raise HTTPException(status_code=403, detail="Access denied")

    schema_ver = db.query(SchemaVersion).filter(
        SchemaVersion.schema_id == source.schema_id
    ).order_by(SchemaVersion.version.desc()).first()

    if not schema_ver:
        return {"fields": [], "name": "", "extraction_instructions": ""}

    defn = schema_ver.definition or {}
    fields = defn.get("fields", [])
    extras_fields = [f["name"] for f in fields if f.get("extras")]
    extras_source = next((f.get("extras_source") for f in fields if f.get("extras") and f.get("extras_source")), None)

    return {
        "name": source.schema.name if source.schema else "",
        "version": schema_ver.version,
        "definition": defn,
        "fields": fields,
        "extraction_instructions": defn.get("extraction_instructions", ""),
        "grouping_key": defn.get("grouping_key", ""),
        "source_website": defn.get("source_website", ""),
        "base_schema": defn.get("base_schema", ""),
        "has_extras": len(extras_fields) > 0,
        "extras_fields": extras_fields,
        "extras_source": extras_source,
    }


# ─── Admin: Reset source status ───────────────────────────────────────────────

@router.post("/{source_id}/reset")
def reset_source(
    source_id: str,
    clear_records: bool = True,
    reason: str | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Admin-only: Reset a source back to 'not_started'.
    Optionally wipe all extracted records (default: True).
    Use this to recover from bad extractions or test data.

    IMPORTANT: extraction_started_at (and the LLM verification timestamps)
    are intentionally NOT cleared. The client-facing timesheet reports total
    elapsed time from when work truly began — a reset because the wrong
    file was uploaded is wasted time that should still count, not a reason
    to make the clock look like it started fresh. The reset itself is
    logged with a reason so it's fully visible in the report.
    """
    source = _get_source_or_404(source_id, db)

    user_roles = {r.role.value for r in current_user.roles}
    if "org_admin" not in user_roles:
        from app.models.all_models import ProjectMember as PM
        member = db.query(PM).filter(PM.project_id == source.project_id, PM.user_id == current_user.id).first()
        if not member or member.role.value not in ("project_admin", "org_admin"):
            raise HTTPException(status_code=403, detail="Only admins can reset sources")

    previous_status = source.status.value
    records_deleted = 0

    if clear_records:
        records_deleted = _safe_clear_old_jobs(source_id, db)

    source.status = SourceStatus.NOT_STARTED
    source.total_records = 0
    source.valid_records = 0
    source.invalid_records = 0
    source.approved_records = 0
    source.review_started_at = None
    source.review_completed_at = None
    source.approved_at = None
    source.web_verified = None
    source.web_check_summary = None
    source.reset_count = (source.reset_count or 0) + 1

    db.add(AuditLog(
        user_id=current_user.id, project_id=source.project_id, source_id=source.id,
        action=AuditAction.SOURCE_RESET,
        before_value={"status": previous_status},
        after_value={
            "status": "not_started",
            "reason": reason or "",
            "records_cleared": clear_records,
            "records_deleted": records_deleted,
            "reset_count": source.reset_count,
            "extraction_timer_preserved": True,
        },
    ))

    db.commit()
    db.refresh(source)

    return {
        "message": f"Source reset to 'not_started' {'with records cleared' if clear_records else 'status only'}",
        "source_id": source_id,
        "status": source.status,
        "records_cleared": clear_records,
        "reset_count": source.reset_count,
    }


# ─── Admin: Clear all records from a source (keep source, keep status) ────────

@router.delete("/{source_id}/records", status_code=200)
def clear_source_records(
    source_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Admin-only: Delete ALL records from a source without changing its status.
    Use this to wipe test data before a real extraction run.
    """
    source = _get_source_or_404(source_id, db)

    if not _can_manage_source(current_user, source):
        raise HTTPException(status_code=403, detail="Only admins can clear records")

    job_ids = [j.id for j in db.query(ExtractionJob).filter(
        ExtractionJob.source_id == source_id
    ).all()]

    deleted_records = 0
    if job_ids:
        deleted_records = db.query(ExtractedRecord).filter(
            ExtractedRecord.job_id.in_(job_ids)
        ).delete(synchronize_session=False)
        db.query(ExtractionJob).filter(
            ExtractionJob.id.in_(job_ids)
        ).delete(synchronize_session=False)

    source.total_records = 0
    source.valid_records = 0
    source.invalid_records = 0
    source.approved_records = 0
    source.web_verified = None
    source.web_check_summary = None
    db.commit()

    return {
        "message": f"Cleared {deleted_records} records from source",
        "records_deleted": deleted_records,
        "source_id": source_id,
    }


# ─── Admin: Unlock submitted records (allow re-review and re-submit) ─────────

@router.post("/{source_id}/unlock", status_code=200)
def unlock_source_records(
    source_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Admin-only: Unlock all submitted records on a source so they can be
    corrected and re-submitted. Resets is_submitted, submitted_at, and
    moves the source status back to in_review.
    """
    source = _get_source_or_404(source_id, db)

    if not _can_manage_source(current_user, source):
        raise HTTPException(status_code=403, detail="Only admins can unlock submitted records")

    job_ids = [
        j.id for j in db.query(ExtractionJob).filter(
            ExtractionJob.source_id == source_id
        ).all()
    ]

    unlocked = 0
    if job_ids:
        records = db.query(ExtractedRecord).filter(
            ExtractedRecord.job_id.in_(job_ids),
            ExtractedRecord.is_submitted == True,
        ).all()
        for r in records:
            r.is_submitted = False
            r.submitted_at = None
            unlocked += 1

        for job in db.query(ExtractionJob).filter(ExtractionJob.id.in_(job_ids)).all():
            job.total_submitted = 0
            job.status = JobStatus.VALIDATED

    if unlocked > 0:
        source.status = SourceStatus.IN_REVIEW

    db.add(AuditLog(
        user_id=current_user.id,
        project_id=source.project_id,
        action=AuditAction.SOURCE_STATUS_CHANGED,
        before_value={"status": "approved", "submitted": True},
        after_value={"status": "in_review", "unlocked_records": unlocked},
    ))
    db.commit()

    return {
        "message": f"Unlocked {unlocked} record(s) — source moved back to In Review",
        "unlocked": unlocked,
        "source_id": source_id,
    }

@router.delete("/{source_id}/records/{record_id}/flags/{flag_index}")
def dismiss_flag(
    source_id: str,
    record_id: str,
    flag_index: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Dismiss one web-check flag by index. Available to reviewers and above.
    The flag is removed permanently from the record — use when the LLM flagged
    something that is actually correct.
    """
    record = db.query(ExtractedRecord).join(
        ExtractionJob, ExtractedRecord.job_id == ExtractionJob.id
    ).filter(
        ExtractedRecord.id == record_id,
        ExtractionJob.source_id == source_id,
    ).first()
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")

    flags = list(record.web_check_flags or [])
    if flag_index < 0 or flag_index >= len(flags):
        raise HTTPException(status_code=400, detail=f"Flag index {flag_index} out of range")

    removed = flags.pop(flag_index)
    record.web_check_flags = flags
    if not flags:
        record.web_verified = True
        record.web_check_summary = "All flags dismissed by reviewer"

    db.commit()
    return {"dismissed": removed, "remaining_flags": len(flags)}


class EscalateNoDataRequest(BaseModel):
    reason: str
    note: str = ""


@router.post("/{source_id}/escalate-no-data")
def escalate_no_data(
    source_id: str,
    payload: EscalateNoDataRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    "Escalate — No Data Found": an extractor's declaration that a source
    has nothing extractable (out of scope, broken page, hoax, etc.), per
    SOP-DS-003 Section 8. Creates a special ExtractedRecord flagged
    is_escalation_only=True, going through the exact same review pipeline
    as a real record — a reviewer still approves or rejects it; rejecting
    sends it to Escalations automatically via the existing mechanism,
    same as any other returned-for-correction record.
    """
    source = _get_source_or_404(source_id, db)
    user_roles = {r.role.value for r in current_user.roles}
    is_admin = "org_admin" in user_roles or "project_admin" in user_roles
    if not is_admin and source.assigned_extractor_id != current_user.id:
        raise HTTPException(status_code=403, detail="Only the assigned extractor or an admin can escalate this source")

    if not payload.reason.strip():
        raise HTTPException(status_code=422, detail="A reason is required")

    job = ExtractionJob(
        project_id=source.project_id, source_id=source.id, schema_id=source.schema_id,
        name=f"{source.name} — Escalation (No Data Found)",
        source_type=FileSourceType.CSV,  # placeholder — not a meaningful classification for an escalation-only job
        status=JobStatus.READY_FOR_REVIEW,
        total_extracted=1, created_by=current_user.id,
    )
    db.add(job)
    db.flush()

    record = ExtractedRecord(
        job_id=job.id,
        extraction_confidence=ExtractionConfidence.HIGH,
        is_schema_valid=True, validation_errors=[],
        review_status=ReviewStatus.PENDING,
        extracted_fields={}, raw_text=payload.note or "(no note provided)",
        is_escalation_only=True, escalation_reason=payload.reason,
    )
    db.add(record)
    db.flush()

    source.status = SourceStatus.READY_FOR_REVIEW
    source.total_records = (source.total_records or 0) + 1

    db.add(AuditLog(
        user_id=current_user.id, project_id=source.project_id, source_id=source.id,
        action=AuditAction.SOURCE_STATUS_CHANGED,
        after_value={"stage": "escalated_no_data", "reason": payload.reason},
    ))
    db.commit()

    return {"job_id": job.id, "record_id": record.id, "status": "ready_for_review"}
